"""
File I/O for the Deleted Showtimes Check batch job: parsing uploads and
building the output workbook (main sheet + SERP_EVIDENCE sheet), ported from
showtime_serp_check.py's load_rows/write_output.

Pure, side-effect-free (no DB/Celery/S3 imports) — unit-testable in isolation,
matching the app/title_matching/batch_io.py convention used elsewhere.
"""

from __future__ import annotations

import csv
import io
from typing import Any, Dict, List, Tuple

import openpyxl
from openpyxl.styles import Font, PatternFill

from app.deleted_showtimes.core import FALSE_, TRUE_, UNKNOWN_, ShowtimeRow
from app.deleted_showtimes.normalize import fmt_min, parse_show_date, parse_time_to_min

REQUIRED_COLUMNS = ("Theater Name", "Title", "Show date", "Show time")
VERDICT_COL = "DELETED_SHOWTIME"
YELLOW = "FFFFFF00"


def peek_headers(contents: bytes, ext: str) -> list[str]:
    ext = ext.lower()
    if ext == ".csv":
        text = contents.decode("utf-8-sig", errors="replace")
        reader = csv.reader(io.StringIO(text))
        raw = next(reader, [])
        return [h.strip() for h in raw]

    wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
    try:
        ws = wb.active
        return [str(ws.cell(1, c).value or "").strip() for c in range(1, (ws.max_column or 0) + 1)]
    finally:
        wb.close()


def parse_upload(contents: bytes, ext: str) -> Tuple[List[str], List[Dict[str, Any]]]:
    """
    Parse an uploaded .csv/.xlsx into (original-case headers, row dicts).

    Validates that all REQUIRED_COLUMNS are present (case-insensitively),
    raising ValueError naming the missing column(s) otherwise. The uploaded
    file must NOT already contain a DELETED_SHOWTIME column — that's the
    output this job produces, not an input.
    """
    ext = ext.lower()
    if ext == ".csv":
        headers, rows = _parse_csv(contents)
    elif ext == ".xlsx":
        headers, rows = _parse_xlsx(contents)
    else:
        raise ValueError(f"Unsupported file extension: {ext!r}")

    lower_headers = {h.strip().lower() for h in headers}
    missing = [col for col in REQUIRED_COLUMNS if col.lower() not in lower_headers]
    if missing:
        raise ValueError(f"Missing required column(s): {', '.join(missing)}")

    return headers, rows


def _parse_csv(contents: bytes) -> Tuple[List[str], List[Dict[str, Any]]]:
    text = contents.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    headers = [h.strip() if h is not None else "" for h in (reader.fieldnames or [])]
    rows = list(reader)
    return headers, rows


def _parse_xlsx(contents: bytes) -> Tuple[List[str], List[Dict[str, Any]]]:
    wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
    try:
        ws = wb.active
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        headers = [str(v).strip() if v is not None else "" for v in header_row]
        rows = []
        for excel_row in ws.iter_rows(min_row=2, values_only=True):
            if all(v is None or str(v).strip() == "" for v in excel_row):
                continue
            row = {}
            for i, v in enumerate(excel_row):
                if i >= len(headers):
                    continue
                if hasattr(v, "hour"):  # datetime.time from an xlsx time-formatted cell
                    row[headers[i]] = v
                else:
                    row[headers[i]] = str(v).strip() if v is not None else ""
            rows.append(row)
        return headers, rows
    finally:
        wb.close()


def rows_to_showtime_rows(headers: List[str], rows: List[Dict[str, Any]]) -> List[ShowtimeRow]:
    """Build ShowtimeRow objects (key = row index into `rows`) for the core
    matching engine, tolerant of case in the required header names."""
    header_map = {h.strip().lower(): h for h in headers}

    def g(row: Dict[str, Any], name: str) -> Any:
        key = header_map.get(name.lower())
        if key is None:
            return ""
        return row.get(key, "")

    out: List[ShowtimeRow] = []
    for idx, row in enumerate(rows):
        t_raw = g(row, "Show time")
        show_min = parse_time_to_min(t_raw)
        show_time_raw = (
            fmt_min(show_min) if hasattr(t_raw, "hour") and show_min is not None
            else str(t_raw).strip()
        )
        out.append(ShowtimeRow(
            key=idx,
            theater=str(g(row, "Theater Name")).strip(),
            title=str(g(row, "Title")).strip(),
            show_date=parse_show_date(g(row, "Show date")),
            show_time_raw=show_time_raw,
            show_min=show_min,
        ))
    return out


def build_output_xlsx(
    original_headers: List[str],
    rows: List[Dict[str, Any]],
    showtime_rows: List[ShowtimeRow],
    no_highlight: bool = False,
) -> bytes:
    """
    Build the output workbook: original rows + a DELETED_SHOWTIME column,
    with TRUE rows highlighted yellow, plus a SERP_EVIDENCE sheet — mirroring
    showtime_serp_check.py's write_output.
    """
    wb = openpyxl.Workbook()
    ws = wb.active

    out_headers = list(original_headers) + [VERDICT_COL]
    ws.append(out_headers)
    for c in ws[1]:
        c.font = Font(bold=True)

    fill = PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type="solid")
    for row, r in zip(rows, showtime_rows):
        out_row = [row.get(h, "") for h in original_headers] + [r.verdict]
        ws.append(out_row)
        if r.verdict == TRUE_ and not no_highlight:
            for c in ws[ws.max_row]:
                c.fill = fill
    ws.column_dimensions[openpyxl.utils.get_column_letter(len(out_headers))].width = 22

    ev = wb.create_sheet("SERP_EVIDENCE")
    ev.append(["Row", "Theater Name", "Title", "Show date", "Show time",
               VERDICT_COL, "REASON", "PUBLISHED_TIMES (SerpApi)", "NEAREST_PUBLISHED",
               "THEATER_VERIFIED", "GOOGLE_THEATER_NAME", "GOOGLE_ADDRESS", "SERPAPI_QUERY"])
    for c in ev[1]:
        c.font = Font(bold=True)
    for row, r in zip(rows, showtime_rows):
        ev.append([
            r.key + 2, r.theater, r.title, str(r.show_date or ""), r.show_time_raw,
            r.verdict, r.reason, r.published, r.nearest,
            "YES" if r.theater_verified else "NO", r.google_theater, r.google_address,
            r.source_query,
        ])
    for i, w in enumerate([8, 34, 30, 12, 11, 22, 44, 60, 20, 16, 38, 40, 40], start=1):
        ev.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ev.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def unknown_row_result(reason: str) -> Tuple[str, str]:
    """Standard (verdict, reason) tuple for a row that couldn't be checked."""
    return UNKNOWN_, reason
