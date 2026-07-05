"""
Parse Amenities Priority.xlsx per §5A spec and seed the database.

Sheet 1 structure:
  - Rows where col A starts with AMENITIES_PRIORITY_N mark the start of a priority tier.
  - Subsequent rows are keyword/format pairs until the next tier marker.

Sheet 3 structure:
  - min_row=2 (row 1 is header)
  - col A: keyword, col B: screen_format, col C: circuit_name
  - circuit_name == "NA" signals a na_default row (not a circuit-specific entry)
  - All other rows become AmenityMapping rows with circuit_name set (P4)
"""

import openpyxl
from sqlalchemy import text as sa_text
from app.detection.normalizer import normalize_string
from app.models import AmenityMapping, CircuitAlias

_VIP_CIRCUIT_ALIASES = [
    "Caribbean Cinemas - US Territory",
    "Cineplex Entertainment",
]


def _clean_cell(value) -> str:
    s = str(value or "").strip()
    s = s.replace("\xa0", " ").replace("xa0", " ")
    s = s.replace("‘", "'").replace("’", "'")
    s = s.replace("“", '"').replace("”", '"')
    return s.strip()


def parse_xlsx(path: str):
    """
    Parse the xlsx file and return (mappings, aliases).

    Sheet 1 → global AmenityMapping rows (no circuit_name).
    Sheet 3 → circuit-specific AmenityMapping rows (circuit_name set, tier=4)
              or na_default attachments when circuit == "NA".
    """
    wb = openpyxl.load_workbook(path, data_only=True)

    # --- Sheet 1: Priority tiers → global mappings ---
    mappings: list[AmenityMapping] = []
    current_tier: int | None = None

    for row in wb.worksheets[0].iter_rows(values_only=True):
        col_a = _clean_cell(row[0] if row else "")
        col_b = _clean_cell(row[1]) if len(row) > 1 and row[1] is not None else ""

        if col_a.startswith("AMENITIES_PRIORITY_"):
            tier_suffix = col_a.replace("AMENITIES_PRIORITY_", "")
            try:
                current_tier = int(tier_suffix)
            except ValueError:
                pass
            continue

        if current_tier is not None and col_a and col_b:
            mappings.append(
                AmenityMapping(
                    amenity_keyword=col_a,
                    screen_format=col_b,
                    priority_tier=current_tier,
                    status="approved",
                )
            )

    # Dedupe within tier
    seen: set[tuple] = set()
    deduped: list[AmenityMapping] = []
    for m in mappings:
        key = (normalize_string(m.amenity_keyword).lower(), m.screen_format, m.priority_tier)
        if key not in seen:
            seen.add(key)
            deduped.append(m)

    # --- Sheet 3: Circuit-specific mappings ---
    alias_map: dict[str, str] = {}

    sheet3 = wb["Sheet3"] if "Sheet3" in wb.sheetnames else None
    if sheet3 is not None:
        for row in sheet3.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue

            kw = _clean_cell(row[0])
            fmt = _clean_cell(row[1]) if len(row) > 1 and row[1] is not None else ""
            circ = _clean_cell(row[2]) if len(row) > 2 and row[2] is not None else ""

            if not kw or not fmt:
                continue

            if circ.upper() == "NA" or not circ:
                # na_default row — attach to matching global mapping
                for m in deduped:
                    if normalize_string(m.amenity_keyword).lower() == normalize_string(kw).lower():
                        m.na_default = fmt
            else:
                # Circuit-specific mapping → AmenityMapping with circuit_name
                deduped.append(
                    AmenityMapping(
                        amenity_keyword=kw,
                        screen_format=fmt,
                        priority_tier=4,
                        circuit_name=circ,
                        status="approved",
                    )
                )
                alias_map[circ.lower()] = circ

    # Build aliases from all circuit names encountered
    aliases: list[CircuitAlias] = [
        CircuitAlias(raw_or_alias=raw, canonical=canonical)
        for raw, canonical in alias_map.items()
    ]

    for canonical in _VIP_CIRCUIT_ALIASES:
        raw = canonical.lower()
        if raw not in alias_map:
            aliases.append(CircuitAlias(raw_or_alias=raw, canonical=canonical))

    return deduped, aliases


def seed_db(session, path: str) -> None:
    mappings, aliases = parse_xlsx(path)

    session.exec(sa_text("TRUNCATE amenitymapping, circuitalias RESTART IDENTITY CASCADE"))
    session.add_all(mappings)
    session.add_all(aliases)
    session.commit()
