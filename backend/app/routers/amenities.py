from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Request
from fastapi.responses import StreamingResponse
from sqlmodel import Session, select
from typing import Optional
import io
import json

import openpyxl

from app.database import get_session
from app.models import AmenityMapping, AuditLog, ReviewItem
from app.schemas import AmenityMappingCreate, AmenityMappingPatch, ReviewDecision
from app.detection.loader import build_engine_from_db

router = APIRouter(prefix="/api/v1/amenities", tags=["amenities"])


def _json_safe(obj):
    return json.dumps(obj, default=str)


def write_audit(
    session: Session,
    table: str,
    record_id,
    action: str,
    before=None,
    after=None,
    actor=None,
) -> None:
    session.add(
        AuditLog(
            table_name=table,
            record_id=str(record_id),
            action=action,
            before_json=_json_safe(before) if before else None,
            after_json=_json_safe(after) if after else None,
            actor=actor,
        )
    )


@router.get("")
def list_amenities(
    search: Optional[str] = None,
    status: Optional[str] = None,
    tier: Optional[str] = None,
    circuit: Optional[str] = None,
    skip: int = 0,
    limit: int = 50,
    session: Session = Depends(get_session),
):
    q = select(AmenityMapping)
    if search:
        q = q.where(AmenityMapping.amenity_keyword.contains(search))
    if status:
        q = q.where(AmenityMapping.status == status)
    if tier:
        tier_int = int(tier.lstrip("P"))
        q = q.where(AmenityMapping.priority_tier == tier_int)
    if circuit:
        q = q.where(AmenityMapping.circuit_name == circuit)
    q = q.offset(skip).limit(limit)
    return session.exec(q).all()


@router.post("")
def create_amenity(
    data: AmenityMappingCreate,
    request: Request,
    session: Session = Depends(get_session),
):
    m = AmenityMapping(**data.dict(), status="pending")
    session.add(m)
    session.flush()
    session.add(
        ReviewItem(
            type="mapping",
            mapping_id=m.id,
            source_string=m.amenity_keyword,
            suggested_format=m.screen_format,
            reasoning=f"Manual submission: {m.amenity_keyword} → {m.screen_format}",
        )
    )
    write_audit(session, "amenity_mappings", m.id, "create", after=data.dict())
    session.commit()
    session.refresh(m)
    return m


@router.put("/{id}")
def update_amenity(
    id: int,
    data: AmenityMappingCreate,
    request: Request,
    session: Session = Depends(get_session),
):
    m = session.get(AmenityMapping, id)
    if not m:
        raise HTTPException(404)
    before = m.dict()
    for k, v in data.dict().items():
        setattr(m, k, v)
    m.status = "pending"
    m.version += 1
    session.add(
        ReviewItem(
            type="mapping",
            mapping_id=m.id,
            source_string=m.amenity_keyword,
            suggested_format=m.screen_format,
            reasoning=f"Manual update: {m.amenity_keyword} → {m.screen_format}",
        )
    )
    write_audit(session, "amenity_mappings", id, "update", before=before, after=data.dict())
    session.commit()
    session.refresh(m)
    return m


@router.patch("/{id}")
def patch_amenity(
    id: int,
    data: AmenityMappingPatch,
    request: Request,
    session: Session = Depends(get_session),
):
    m = session.get(AmenityMapping, id)
    if not m:
        raise HTTPException(404)
    before = m.dict()
    patch_data = data.dict(exclude_unset=True)
    for k, v in patch_data.items():
        setattr(m, k, v)
    if "status" not in patch_data:
        m.status = "pending"
    m.version += 1
    session.add(
        ReviewItem(
            type="mapping",
            mapping_id=m.id,
            source_string=m.amenity_keyword,
            suggested_format=m.screen_format,
            reasoning=f"Manual update: {m.amenity_keyword} → {m.screen_format}",
        )
    )
    write_audit(session, "amenity_mappings", id, "patch", before=before, after=patch_data)
    session.commit()
    session.refresh(m)
    return m


@router.post("/{id}/approve")
def approve_amenity(
    id: int,
    request: Request,
    session: Session = Depends(get_session),
):
    m = session.get(AmenityMapping, id)
    if not m:
        raise HTTPException(404)
    before = {"status": m.status}
    m.status = "approved"
    write_audit(
        session,
        "amenity_mappings",
        id,
        "approve",
        before=before,
        after={"status": "approved"},
    )
    session.commit()
    request.app.state.engine = build_engine_from_db(session)
    return {"ok": True}


@router.post("/{id}/reject")
def reject_amenity(
    id: int,
    body: ReviewDecision,
    request: Request,
    session: Session = Depends(get_session),
):
    m = session.get(AmenityMapping, id)
    if not m:
        raise HTTPException(404)
    before = {"status": m.status}
    m.status = "rejected"
    write_audit(
        session,
        "amenity_mappings",
        id,
        "reject",
        before=before,
        after={"status": "rejected", "reason": body.reason},
    )
    session.commit()
    return {"ok": True}


@router.post("/import")
async def import_xlsx(
    file: UploadFile = File(...),
    session: Session = Depends(get_session),
):
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active
    headers = [
        str(ws.cell(1, c).value or "").strip().lower()
        for c in range(1, ws.max_column + 1)
    ]
    required = ["amenity_keyword", "screen_format", "priority_tier"]
    for r in required:
        if r not in headers:
            raise HTTPException(400, detail=f"Column '{r}' not found")
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = {headers[i]: row[i] for i in range(len(headers)) if i < len(row)}
        if not row_dict.get("amenity_keyword") or not row_dict.get("screen_format"):
            continue
        m = AmenityMapping(
            amenity_keyword=str(row_dict["amenity_keyword"]),
            screen_format=str(row_dict["screen_format"]),
            priority_tier=int(row_dict.get("priority_tier") or 4),
            status="pending",
        )
        session.add(m)
        count += 1
    session.commit()
    return {"imported": count}


@router.get("/export")
def export_xlsx(session: Session = Depends(get_session)):
    mappings = session.exec(
        select(AmenityMapping).where(AmenityMapping.status == "approved")
    ).all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(
        ["amenity_keyword", "screen_format", "priority_tier", "circuit_name", "na_default", "status"]
    )
    for m in mappings:
        ws.append(
            [
                m.amenity_keyword,
                m.screen_format,
                m.priority_tier,
                m.circuit_name or "",
                m.na_default or "",
                m.status,
            ]
        )
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=amenities_export.xlsx"},
    )
