import os
import threading
import uuid

from fastapi import APIRouter, Depends, Form, HTTPException, Query, Request, UploadFile, File
from pydantic import BaseModel
from sqlmodel import Session, select
from typing import Optional

from app.config import settings
from app.database import get_session
from app.detection.bedrock_client import bedrock_client
from app.models import AmenityMapping, DetectionJob, ReviewItem

router = APIRouter(prefix="/api/v1/detect", tags=["detect"])

_UPLOAD_DIR = "/tmp/amenity_uploads"


class DetectSingleRequest(BaseModel):
    amenity: str
    circuit_name: Optional[str] = ""


@router.post("/single")
async def detect_single(
    payload: DetectSingleRequest,
    request: Request,
    session: Session = Depends(get_session),
):
    engine = request.app.state.engine
    result = engine.detect(payload.amenity, payload.circuit_name or "")

    if result.fired_ai and settings.AI_TRIGGER_MODE != "off":
        known_formats = sorted({
            m.screen_format
            for m in session.exec(
                select(AmenityMapping).where(AmenityMapping.status == "approved")
            ).all()
        })
        suggestion = bedrock_client.classify_single(
            amenity=payload.amenity,
            circuit=payload.circuit_name or "",
            known_formats=known_formats,
        )
        if suggestion:
            result.ai_suggested_format = suggestion.suggested_screen_format
            result.ai_reasoning = suggestion.reasoning

            # Push to review queue so a human can approve → becomes a mapping
            session.add(ReviewItem(
                type="ai_suggestion",
                source_string=payload.amenity,
                circuit=payload.circuit_name or None,
                suggested_format=suggestion.suggested_screen_format,
                confidence=suggestion.confidence,
                reasoning=suggestion.reasoning,
            ))
            session.commit()

    return result.__dict__ if hasattr(result, "__dict__") else result


@router.post("/batch")
async def detect_batch(
    request: Request,
    file: UploadFile = File(...),
    include_diagnostics: str = Form("false"),
    audit_mode: bool = Query(False),
    session: Session = Depends(get_session),
):
    diag_bool = include_diagnostics.lower() in ("true", "1", "yes")
    filename = file.filename or ""
    ext = os.path.splitext(filename)[1].lower()
    if ext not in (".xlsx", ".csv"):
        raise HTTPException(400, detail="Only .xlsx and .csv files are supported")

    contents = await file.read()
    row_count = _estimate_rows(contents, ext)
    if row_count > settings.MAX_BATCH_ROWS:
        raise HTTPException(400, detail=f"File exceeds {settings.MAX_BATCH_ROWS} row limit")

    # Validate required columns before accepting the job
    from app.workers.batch_worker import _peek_headers
    headers = _peek_headers(contents, ext)
    required_cols = ["amenities", "circuit_name"]
    if audit_mode:
        required_cols.append("screen_format")
    missing = [col for col in required_cols if col not in headers]
    if missing:
        raise HTTPException(400, detail=f"Missing required column(s): {', '.join(missing)}")

    os.makedirs(_UPLOAD_DIR, exist_ok=True)
    job_id = str(uuid.uuid4())
    upload_path = os.path.join(_UPLOAD_DIR, f"{job_id}{ext}")
    with open(upload_path, "wb") as f_out:
        f_out.write(contents)

    job = DetectionJob(
        id=job_id,
        status="queued",
        total=row_count,
        include_diagnostics=diag_bool,
        audit_mode=audit_mode,
    )
    session.add(job)
    session.commit()

    from app.workers.batch_worker import run_batch_job
    t = threading.Thread(
        target=run_batch_job,
        args=(job_id, upload_path, diag_bool, request.app.state.engine),
        kwargs={"audit_mode": audit_mode},
        daemon=True,
    )
    t.start()

    return {"job_id": job_id}


def _estimate_rows(contents: bytes, ext: str) -> int:
    """Fast row count without full parse — used for limit check."""
    if ext == ".csv":
        import io as _io
        text = contents.decode("utf-8-sig", errors="replace")
        return max(0, text.count("\n") - 1)
    # xlsx: load header only for speed
    try:
        import openpyxl
        import io as _io
        wb = openpyxl.load_workbook(_io.BytesIO(contents), read_only=True, data_only=True)
        ws = wb.active
        return max(0, ws.max_row - 1) if ws.max_row else 0
    except Exception:
        return 0
