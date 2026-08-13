from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Request
from fastapi.responses import StreamingResponse
from sqlmodel import Session, select, func
from typing import Optional
import io
import json
import math

import openpyxl

from app.database import get_session
from app.models import IntlAmenityMapping, AuditLog
from app.schemas import (
    IntlAmenityMappingCreate,
    IntlAmenityMappingPatch,
    ReviewDecision,
    PaginatedResponse,
)
from app.intl_detection.loader import build_intl_engine_from_db

router = APIRouter(prefix="/api/v1/intl-amenities", tags=["intl-amenities"])


def _json_safe(obj):
    return json.dumps(obj, default=str)


def write_audit(
    session: Session,
    table: str,
    record_id,
    action: str,
    before=None,
    after=None,
    actor=None,
) -> None:
    session.add(
        AuditLog(
            table_name=table,
            record_id=str(record_id),
            action=action,
            before_json=_json_safe(before) if before else None,
            after_json=_json_safe(after) if after else None,
            actor=actor,
        )
    )


@router.get("")
def list_intl_amenities(
    search: Optional[str] = None,
    status: Optional[str] = None,
    tier: Optional[str] = None,
    page: int = 1,
    page_size: int = 50,
    session: Session = Depends(get_session),
):
    q = select(IntlAmenityMapping)
    if search:
        q = q.where(IntlAmenityMapping.amenity_keyword.contains(search))
    if status:
        q = q.where(IntlAmenityMapping.status == status)
    if tier:
        tier_int = int(tier.lstrip("P"))
        q = q.where(IntlAmenityMapping.priority_tier == tier_int)
    # NOTE: no `circuit` filter — that column is dormant for intl, and an
    # always-empty filter would be a UI liability.

    count_q = select(func.count()).select_from(q.subquery())
    total = session.exec(count_q).one()

    q = q.order_by(IntlAmenityMapping.priority_tier, IntlAmenityMapping.amenity_keyword)
    items = session.exec(q.offset((page - 1) * page_size).limit(page_size)).all()
    return PaginatedResponse(
        items=items,
        total=total,
        page=page,
        page_size=page_size,
        total_pages=max(1, math.ceil(total / page_size)),
    )


@router.post("")
def create_intl_amenity(
    data: IntlAmenityMappingCreate,
    request: Request,
    session: Session = Depends(get_session),
):
    incoming = data.dict()
    m = IntlAmenityMapping(**incoming, status="pending")
    session.add(m)
    write_audit(session, "intlamenitymapping", m.id, "create", after=incoming)
    session.commit()
    session.refresh(m)
    # status is always "pending" on create, so the approved set never
    # changes here — no engine rebuild needed. (approve/reject/patch/delete
    # below are the mutations that can change it.)
    return m


@router.patch("/{id}")
def patch_intl_amenity(
    id: int,
    data: IntlAmenityMappingPatch,
    request: Request,
    session: Session = Depends(get_session),
):
    m = session.get(IntlAmenityMapping, id)
    if not m:
        raise HTTPException(404)
    before = m.dict()
    patch_data = data.dict(exclude_unset=True)
    for k, v in patch_data.items():
        setattr(m, k, v)
    if patch_data.get("status") != "approved":
        m.status = "pending"
    m.version += 1
    write_audit(session, "intlamenitymapping", id, "patch", before=before, after=patch_data)
    session.commit()
    session.refresh(m)
    request.app.state.intl_engine = build_intl_engine_from_db(session)
    return m


@router.post("/{id}/approve")
def approve_intl_amenity(
    id: int,
    request: Request,
    session: Session = Depends(get_session),
):
    m = session.get(IntlAmenityMapping, id)
    if not m:
        raise HTTPException(404)
    before = {"status": m.status}
    m.status = "approved"
    write_audit(
        session,
        "intlamenitymapping",
        id,
        "approve",
        before=before,
        after={"status": "approved"},
    )
    session.commit()
    request.app.state.intl_engine = build_intl_engine_from_db(session)
    return {"ok": True}


@router.post("/{id}/reject")
def reject_intl_amenity(
    id: int,
    body: ReviewDecision,
    request: Request,
    session: Session = Depends(get_session),
):
    m = session.get(IntlAmenityMapping, id)
    if not m:
        raise HTTPException(404)
    before = {"status": m.status}
    m.status = "rejected"
    write_audit(
        session,
        "intlamenitymapping",
        id,
        "reject",
        before=before,
        after={"status": "rejected", "reason": body.reason},
    )
    session.commit()
    request.app.state.intl_engine = build_intl_engine_from_db(session)
    return {"ok": True}


@router.delete("/{id}")
def delete_intl_amenity(
    id: int,
    request: Request,
    session: Session = Depends(get_session),
):
    m = session.get(IntlAmenityMapping, id)
    if not m:
        raise HTTPException(404)
    write_audit(session, "intlamenitymapping", id, "delete", before=m.dict())
    session.delete(m)
    session.commit()
    request.app.state.intl_engine = build_intl_engine_from_db(session)
    return {"ok": True}


@router.post("/import")
async def import_intl_xlsx(
    file: UploadFile = File(...),
    session: Session = Depends(get_session),
):
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active
    headers = [
        str(ws.cell(1, c).value or "").strip().lower()
        for c in range(1, ws.max_column + 1)
    ]
    required = ["amenity_keyword", "screen_format", "priority_tier"]
    for r in required:
        if r not in headers:
            raise HTTPException(400, detail=f"Column '{r}' not found")
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = {headers[i]: row[i] for i in range(len(headers)) if i < len(row)}
        if not row_dict.get("amenity_keyword") or not row_dict.get("screen_format"):
            continue
        m = IntlAmenityMapping(
            amenity_keyword=str(row_dict["amenity_keyword"]),
            screen_format=str(row_dict["screen_format"]),
            priority_tier=int(row_dict.get("priority_tier") or 4),
            status="pending",
        )
        session.add(m)
        count += 1
    session.commit()
    return {"imported": count}


@router.get("/export")
def export_intl_xlsx(session: Session = Depends(get_session)):
    mappings = session.exec(
        select(IntlAmenityMapping).where(IntlAmenityMapping.status == "approved")
    ).all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["amenity_keyword", "screen_format", "priority_tier", "status"])
    for m in mappings:
        ws.append(
            [
                m.amenity_keyword,
                m.screen_format,
                m.priority_tier,
                m.status,
            ]
        )
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=intl_amenities_export.xlsx"},
    )
