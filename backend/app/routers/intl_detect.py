import os
import threading
import uuid

from fastapi import APIRouter, Depends, Form, HTTPException, Query, Request, UploadFile, File
from pydantic import BaseModel
from sqlmodel import Session

from app.config import settings
from app.database import get_session
from app.models import IntlDetectionJob

router = APIRouter(prefix="/api/v1/intl-detect", tags=["intl-detect"])

_UPLOAD_DIR = "/tmp/intl_amenity_uploads"


class IntlDetectSingleRequest(BaseModel):
    amenity: str


@router.post("/single")
async def detect_single_intl(
    payload: IntlDetectSingleRequest,
    request: Request,
):
    engine = request.app.state.intl_engine
    result = engine.detect(payload.amenity)
    return result.__dict__ if hasattr(result, "__dict__") else result


@router.post("/batch")
async def detect_batch_intl(
    request: Request,
    file: UploadFile = File(...),
    include_diagnostics: str = Form("false"),
    audit_mode: bool = Query(False),
    session: Session = Depends(get_session),
):
    diag_bool = include_diagnostics.lower() in ("true", "1", "yes")
    filename = file.filename or ""
    ext = os.path.splitext(filename)[1].lower()
    if ext not in (".xlsx", ".csv"):
        raise HTTPException(400, detail="Only .xlsx and .csv files are supported")

    contents = await file.read()
    row_count = _estimate_rows(contents, ext)
    if row_count > settings.MAX_BATCH_ROWS:
        raise HTTPException(400, detail=f"File exceeds {settings.MAX_BATCH_ROWS} row limit")

    from app.workers.intl_batch_worker import _peek_headers
    headers = _peek_headers(contents, ext)
    has_amenities = "amenities_string" in headers or "amenities" in headers
    if not has_amenities:
        raise HTTPException(400, detail="Missing required column: amenities or amenities_string")
    if audit_mode and "screen_format" not in headers:
        raise HTTPException(400, detail="audit_mode requires column: screen_format")

    # NOTE: circuit_name is deliberately NOT a required upload column here,
    # unlike detect.py's domestic batch endpoint. There is no intl circuit
    # data, so do not "fix" this by adding it back.

    os.makedirs(_UPLOAD_DIR, exist_ok=True)
    job_id = str(uuid.uuid4())
    upload_path = os.path.join(_UPLOAD_DIR, f"{job_id}{ext}")
    with open(upload_path, "wb") as f_out:
        f_out.write(contents)

    job = IntlDetectionJob(
        id=job_id,
        status="queued",
        total=row_count,
        include_diagnostics=diag_bool,
        audit_mode=audit_mode,
    )
    session.add(job)
    session.commit()

    from app.workers.intl_batch_worker import run_intl_batch_job
    t = threading.Thread(
        target=run_intl_batch_job,
        args=(job_id, upload_path, diag_bool, request.app.state.intl_engine),
        kwargs={"audit_mode": audit_mode},
        daemon=True,
    )
    t.start()

    return {"job_id": job_id}


def _estimate_rows(contents: bytes, ext: str) -> int:
    """Copied verbatim from movie_detect.py."""
    if ext == ".csv":
        import io as _io
        text = contents.decode("utf-8-sig", errors="replace")
        return max(0, text.count("\n") - 1)
    try:
        import openpyxl
        import io as _io
        wb = openpyxl.load_workbook(_io.BytesIO(contents), read_only=True, data_only=True)
        ws = wb.active
        return max(0, ws.max_row - 1) if ws.max_row else 0
    except Exception:
        return 0
