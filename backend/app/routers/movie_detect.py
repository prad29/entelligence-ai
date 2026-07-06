import os
import threading
import uuid

from fastapi import APIRouter, Depends, Form, HTTPException, Query, Request, UploadFile, File
from pydantic import BaseModel
from sqlmodel import Session
from typing import Optional

from app.config import settings
from app.database import get_session
from app.detection.bedrock_client import bedrock_client
from app.models import MovieFormatJob, MovieFormatReviewItem

router = APIRouter(prefix="/api/v1/movie-detect", tags=["movie-detect"])

_UPLOAD_DIR = "/tmp/movie_uploads"

_KNOWN_FORMATS = ["70MM", "35MM", "3D", "2D"]


class MovieDetectSingleRequest(BaseModel):
    amenity: str


@router.post("/single")
async def detect_single_movie(
    payload: MovieDetectSingleRequest,
    request: Request,
    session: Session = Depends(get_session),
):
    engine = request.app.state.movie_engine
    result = engine.detect(payload.amenity)

    if result.fired_ai and settings.AI_TRIGGER_MODE != "off":
        from app.cache import get_redis, movie_format_cache_key
        import json as _json

        cache_key = movie_format_cache_key(payload.amenity)
        suggestion = None

        try:
            redis = get_redis()
            cached = redis.get(cache_key)
            if cached:
                from app.movie_detection.types import MovieFormatBedrockSuggestion
                data = _json.loads(cached)
                suggestion = MovieFormatBedrockSuggestion(**data)
        except Exception:
            redis = None

        if suggestion is None:
            suggestion = bedrock_client.classify_single(payload.amenity, "", _KNOWN_FORMATS)
            if suggestion and redis:
                try:
                    ttl = settings.BEDROCK_CACHE_TTL_DAYS * 86400
                    redis.setex(cache_key, ttl, _json.dumps(suggestion.__dict__))
                except Exception:
                    pass

        if suggestion:
            result.ai_suggested_format = suggestion.suggested_screen_format
            result.ai_reasoning = suggestion.reasoning

            session.add(MovieFormatReviewItem(
                type="ai_suggestion",
                source_string=payload.amenity,
                suggested_format=suggestion.suggested_screen_format,
                confidence=suggestion.confidence,
                reasoning=suggestion.reasoning,
            ))
            session.commit()

    return result.__dict__ if hasattr(result, "__dict__") else result


@router.post("/batch")
async def detect_batch_movie(
    request: Request,
    file: UploadFile = File(...),
    include_diagnostics: str = Form("false"),
    batch_ai_mode: str = Form("skip"),
    audit_mode: bool = Query(False),
    session: Session = Depends(get_session),
):
    diag_bool = include_diagnostics.lower() in ("true", "1", "yes")
    filename = file.filename or ""
    ext = os.path.splitext(filename)[1].lower()
    if ext not in (".xlsx", ".csv"):
        raise HTTPException(400, detail="Only .xlsx and .csv files are supported")

    contents = await file.read()
    row_count = _estimate_rows(contents, ext)
    if row_count > settings.MAX_BATCH_ROWS:
        raise HTTPException(400, detail=f"File exceeds {settings.MAX_BATCH_ROWS} row limit")

    from app.workers.movie_batch_worker import _peek_headers
    headers = _peek_headers(contents, ext)
    if audit_mode:
        missing = [col for col in ("amenities_string", "movie_format") if col not in headers]
        if missing:
            raise HTTPException(400, detail=f"audit_mode requires columns: {', '.join(missing)}")
    else:
        if "amenities_string" not in headers and "amenities" not in headers:
            raise HTTPException(400, detail="Missing required column: amenities_string")

    os.makedirs(_UPLOAD_DIR, exist_ok=True)
    job_id = str(uuid.uuid4())
    upload_path = os.path.join(_UPLOAD_DIR, f"{job_id}{ext}")
    with open(upload_path, "wb") as f_out:
        f_out.write(contents)

    job = MovieFormatJob(
        id=job_id,
        status="queued",
        total=row_count,
        include_diagnostics=diag_bool,
        audit_mode=audit_mode,
    )
    session.add(job)
    session.commit()

    from app.workers.movie_batch_worker import run_movie_batch_job
    t = threading.Thread(
        target=run_movie_batch_job,
        args=(job_id, upload_path, diag_bool, request.app.state.movie_engine, batch_ai_mode, audit_mode),
        daemon=True,
    )
    t.start()

    return {"job_id": job_id}


def _estimate_rows(contents: bytes, ext: str) -> int:
    if ext == ".csv":
        import io as _io
        text = contents.decode("utf-8-sig", errors="replace")
        return max(0, text.count("\n") - 1)
    try:
        import openpyxl
        import io as _io
        wb = openpyxl.load_workbook(_io.BytesIO(contents), read_only=True, data_only=True)
        ws = wb.active
        return max(0, ws.max_row - 1) if ws.max_row else 0
    except Exception:
        return 0
