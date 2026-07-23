import csv
import io
import os
import uuid
from datetime import datetime
from typing import Literal, Optional

from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, Request, UploadFile
from fastapi.responses import Response
from pydantic import BaseModel, model_validator
from sqlmodel import Session

from app.config import settings
from app.database import get_session

router = APIRouter(prefix="/api/v1/movie-title-match", tags=["movie-title-match"])


class TitleMatchRequest(BaseModel):
    title: str
    theater: Optional[str] = None
    show_date: Optional[str] = None       # YYYY-MM-DD
    ticketing_url: Optional[str] = None
    use_poster_vision: bool = False       # Mode B only: enable Claude vision on DB posters
    market: Literal["domestic", "international"] = "domestic"
    country: Optional[str] = None         # required when market == "international"

    @model_validator(mode="after")
    def _require_country_for_international(self) -> "TitleMatchRequest":
        if self.market == "international" and not (self.country or "").strip():
            raise ValueError("country is required when market is 'international'")
        return self


@router.post("/single")
async def match_single_title(
    payload: TitleMatchRequest,
    request: Request,
    session: Session = Depends(get_session),
):
    engine = getattr(request.app.state, 'title_match_engine', None)
    if engine is None:
        return {
            "suggested_movie_id": 0,
            "suggested_movie_title": "Engine not loaded",
            "canonical_movie_id": 0,
            "confidence": 0.0,
            "decision": "REVIEW",
            "reasoning": (
                "Movie Master has not been seeded yet. "
                "Run: python app/cli.py seed-movie-master /path/to/dump.csv"
            ),
            "evidence": {},
            "cover_image": None,
            "ticketing_poster_url": None,
            "fired_ai": False,
        }

    result = engine.match(
        title=payload.title,
        show_date=payload.show_date,
        theater=payload.theater,
        ticketing_url=payload.ticketing_url,
        use_poster_vision=payload.use_poster_vision,
        market=payload.market,
        country=payload.country,
    )
    return result.__dict__


@router.post("/batch")
async def upload_batch(
    file: UploadFile = File(...),
    use_poster_vision: str = Form("false"),
    session: Session = Depends(get_session),
):
    """Upload a .csv/.xlsx of titles for Mode B agentic batch matching.

    Requires settings.AGENTIC_TITLE_MATCH_ENABLED (batch matching is Mode B
    only). Validates the file extension, the 3 required columns, and the
    MAX_BATCH_ROWS cap, then enqueues dispatch_batch_task (see
    app.tasks.agentic_match_task) instead of calling dispatch_batch inline.
    Building the chord re-parses the file and publishes one Celery message
    per row — for large files that alone can exceed the ALB/nginx idle
    timeout if done inside the request, so it always runs in the background.
    """
    if not settings.AGENTIC_TITLE_MATCH_ENABLED:
        raise HTTPException(
            status_code=400,
            detail="Batch title matching requires Mode B (agentic) to be enabled",
        )

    from app.models import MovieTitleBatchJob
    from app.tasks.agentic_match_task import dispatch_batch_task
    from app.title_matching import batch_io, batch_storage

    filename = file.filename or ""
    ext = os.path.splitext(filename)[1].lower()
    if ext not in (".csv", ".xlsx"):
        raise HTTPException(status_code=400, detail="Only .csv and .xlsx files are supported")

    contents = await file.read()

    try:
        _headers, rows = batch_io.parse_upload(contents, ext)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    row_count = len(rows)
    if row_count > settings.MAX_BATCH_ROWS:
        raise HTTPException(
            status_code=400,
            detail=f"File exceeds {settings.MAX_BATCH_ROWS} row limit",
        )

    use_poster_vision_bool = use_poster_vision.strip().lower() in ("true", "1", "yes")

    job_id = str(uuid.uuid4())
    upload_key = batch_storage.upload_key(job_id, ext)
    batch_storage.put_bytes(upload_key, contents)

    job = MovieTitleBatchJob(
        id=job_id,
        status="queued",
        total=row_count,
        use_poster_vision=use_poster_vision_bool,
        file_path=upload_key,
    )
    session.add(job)
    session.commit()

    dispatch_batch_task.delay(job_id)

    return {"job_id": job_id}


@router.get("/batch/{job_id}")
async def get_batch_job(job_id: str, session: Session = Depends(get_session)):
    from app.models import MovieTitleBatchJob

    job = session.get(MovieTitleBatchJob, job_id)
    if job is None:
        raise HTTPException(status_code=404, detail="Job not found")

    progress = (job.processed / job.total) if job.total > 0 else 0

    return {
        "job_id": job.id,
        "status": job.status,
        "total": job.total,
        "processed": job.processed,
        "progress": progress,
        "matched": job.matched,
        "no_match": job.no_match,
        "failed": job.failed,
        "output_url": (
            f"/api/v1/movie-title-match/batch/{job.id}/download"
            if job.status == "completed" and job.output_path
            else None
        ),
        "error": job.error,
    }


@router.get("/batch/{job_id}/download")
async def download_batch_job(job_id: str, session: Session = Depends(get_session)) -> Response:
    from app.models import MovieTitleBatchJob
    from app.title_matching import batch_storage

    job = session.get(MovieTitleBatchJob, job_id)
    if job is None:
        raise HTTPException(status_code=404, detail="Job not found")

    if job.status != "completed":
        raise HTTPException(status_code=400, detail="Job not completed")

    if job.ttl and datetime.utcnow() > job.ttl:
        raise HTTPException(status_code=410, detail="Download expired")

    if not job.output_path or not batch_storage.exists(job.output_path):
        raise HTTPException(status_code=404, detail="Output file not found")

    contents = batch_storage.get_bytes(job.output_path)

    return Response(
        content=contents,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="movie_title_match_results_{job_id[:8]}.xlsx"'
        },
    )


@router.get("/master")
async def search_master(
    q: str = Query(""),
    limit: int = Query(20, le=100),
    session: Session = Depends(get_session),
):
    from sqlmodel import select
    from app.models import MovieMaster

    stmt = select(MovieMaster).where(
        MovieMaster.movie_title.ilike(f"%{q}%")
    ).limit(limit)
    rows = session.exec(stmt).all()
    return [
        {
            "id": r.id,
            "movie_title": r.movie_title,
            "release_date": r.release_date,
            "cover_image": r.cover_image,
            "imdb_id": r.imdb_id,
        }
        for r in rows
    ]


@router.get("/master/search")
async def agent_search_master(
    q: str = Query(""),
    limit: int = Query(10, le=50),
    session: Session = Depends(get_session),
):
    """Internal keyword search endpoint used by the Mode B agentic subprocess.

    Returns richer metadata (director, running_time) to help the agent disambiguate.
    No auth — intended for localhost-only use by the claude subprocess.
    """
    from sqlmodel import select
    from app.models import MovieMaster

    stmt = select(MovieMaster).where(
        MovieMaster.movie_title.ilike(f"%{q}%")
    ).limit(limit)
    rows = session.exec(stmt).all()
    return [
        {
            "id": r.id,
            "movie_title": r.movie_title,
            "release_date": r.release_date,
            "imdb_id": r.imdb_id,
            "director": r.director,
            "running_time": r.running_time,
        }
        for r in rows
    ]


@router.get("/master/count")
async def get_master_count(session: Session = Depends(get_session)):
    from sqlmodel import select, func
    from app.models import MovieMaster

    count = session.exec(select(func.count()).select_from(MovieMaster)).one()
    return {"count": count}


@router.post("/master/seed")
async def seed_master(
    file: UploadFile = File(...),
    request: Request = None,
    session: Session = Depends(get_session),
):
    from app.models import MovieMaster
    from sqlmodel import select, func
    from app.title_matching.seed_loader import seed_from_rows

    filename = file.filename or ""
    content = await file.read()

    if filename.endswith(".xlsx") or filename.endswith(".xls"):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        headers = [str(cell.value).strip() if cell.value is not None else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        rows = []
        for excel_row in ws.iter_rows(min_row=2, values_only=True):
            rows.append({headers[i]: (str(v).strip() if v is not None else "") for i, v in enumerate(excel_row)})
        wb.close()
    else:
        text = content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        rows = list(reader)

    existing_count = session.exec(select(func.count()).select_from(MovieMaster)).one()
    result = seed_from_rows(session, rows)

    # Reload the title match engine in app state
    if request is not None:
        try:
            from app.title_matching.loader import build_title_match_engine
            from app.title_matching.engine import TitleMatchEngine
            gen, aliases = build_title_match_engine(session)
            request.app.state.title_match_engine = TitleMatchEngine(gen, aliases)
        except Exception:
            pass

    # Queue semantic index build whenever rows were inserted or updated
    if result["inserted"] > 0 or result["updated"] > 0:
        try:
            from app.tasks.semantic_tasks import build_semantic_index_task
            build_semantic_index_task.delay()
        except Exception:
            pass

    return {
        "previously_seeded": existing_count,
        "inserted": result["inserted"],
        "updated": result["updated"],
        "skipped": result["skipped"],
        "total_in_file": len(rows),
    }


@router.get("/master/intl/count")
async def get_master_intl_count(session: Session = Depends(get_session)):
    from sqlmodel import select, func
    from app.models import MovieMasterIntl

    count = session.exec(select(func.count()).select_from(MovieMasterIntl)).one()
    return {"count": count}


@router.post("/master/intl/seed")
async def seed_master_intl(
    file: UploadFile = File(...),
    session: Session = Depends(get_session),
):
    """Seed the international master table from an uploaded CSV/xlsx.

    Grain is (movie_id, country, release_date) — separate upsert path from
    the domestic /master/seed, which keys on id alone.
    """
    from sqlmodel import select, func
    from app.models import MovieMasterIntl
    from app.title_matching.seed_loader import seed_intl_from_rows

    filename = file.filename or ""
    content = await file.read()

    if filename.endswith(".xlsx") or filename.endswith(".xls"):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        headers = [str(cell.value).strip() if cell.value is not None else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        rows = []
        for excel_row in ws.iter_rows(min_row=2, values_only=True):
            rows.append({headers[i]: (str(v).strip() if v is not None else "") for i, v in enumerate(excel_row)})
        wb.close()
    else:
        text = content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        rows = list(reader)

    existing_count = session.exec(select(func.count()).select_from(MovieMasterIntl)).one()
    result = seed_intl_from_rows(session, rows)

    # Queue the international semantic index build whenever rows changed —
    # runs independently of the domestic build_semantic_index_task.
    if result["inserted"] > 0 or result["updated"] > 0:
        try:
            from app.tasks.semantic_tasks import build_semantic_index_intl_task
            build_semantic_index_intl_task.delay()
        except Exception:
            pass

    return {
        "previously_seeded": existing_count,
        "inserted": result["inserted"],
        "updated": result["updated"],
        "skipped": result["skipped"],
        "skipped_undefined_country": result["skipped_undefined_country"],
        "total_in_file": len(rows),
    }
