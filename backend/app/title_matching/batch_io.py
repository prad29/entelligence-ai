"""
Pure, side-effect-free helpers for batch title-matching file I/O.

No DB, network, or Celery imports at module load time — these functions are
unit-testable in isolation and are shared by the batch upload router and the
Celery task that processes rows.
"""

import csv
import io
from typing import Any, Callable, Optional

import openpyxl

REQUIRED_COLUMNS: tuple[str, str, str] = ("movie_title", "show_date", "ticketing_url")

# "title" is accepted as an alias for "movie_title" — some upstream exports
# (e.g. Tableau mapping lists) use "title" instead of the canonical column name.
TITLE_COLUMN_ALIASES: tuple[str, ...] = ("movie_title", "title")


def peek_headers(contents: bytes, ext: str) -> list[str]:
    """
    Return the original-case header row from raw file bytes without persisting
    anything to disk.

    ext must be either ".csv" or ".xlsx" (case-insensitive).
    """
    ext = ext.lower()
    if ext == ".csv":
        text = contents.decode("utf-8-sig", errors="replace")
        reader = csv.reader(io.StringIO(text))
        raw = next(reader, [])
        return [h.strip() for h in raw]

    wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
    try:
        ws = wb.active
        return [str(ws.cell(1, c).value or "").strip() for c in range(1, (ws.max_column or 0) + 1)]
    finally:
        wb.close()


def parse_upload(contents: bytes, ext: str) -> tuple[list[str], list[dict[str, Any]]]:
    """
    Parse an uploaded .csv or .xlsx file into (original-case headers, row dicts).

    Row dicts are keyed by the original-case header string as it appeared in
    the file. Validates (case-insensitively) that all of REQUIRED_COLUMNS are
    present, raising ValueError naming the missing column(s) if not.
    """
    ext = ext.lower()
    if ext == ".csv":
        headers, rows = _parse_csv(contents)
    elif ext == ".xlsx":
        headers, rows = _parse_xlsx(contents)
    else:
        raise ValueError(f"Unsupported file extension: {ext!r}")

    lower_headers = {h.strip().lower() for h in headers}
    missing = [
        col for col in REQUIRED_COLUMNS
        if col not in lower_headers
        and not (col == "movie_title" and lower_headers & set(TITLE_COLUMN_ALIASES))
    ]
    if missing:
        raise ValueError(
            f"Missing required column(s): {', '.join(missing)}"
        )

    return headers, rows


def _parse_csv(contents: bytes) -> tuple[list[str], list[dict[str, Any]]]:
    text = contents.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    headers = [h.strip() if h is not None else "" for h in (reader.fieldnames or [])]
    rows = list(reader)
    return headers, rows


def _parse_xlsx(contents: bytes) -> tuple[list[str], list[dict[str, Any]]]:
    # openpyxl's max_row reflects the sheet's stored <dimension> metadata, which
    # can be stale (larger than the actual populated range) after rows are
    # deleted in some editors. Skip fully-blank rows rather than trusting
    # max_row, to avoid manufacturing phantom empty-title rows.
    wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
    try:
        ws = wb.active
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        headers = [str(v).strip() if v is not None else "" for v in header_row]
        rows = []
        for excel_row in ws.iter_rows(min_row=2, values_only=True):
            if all(v is None or str(v).strip() == "" for v in excel_row):
                continue
            rows.append(
                {headers[i]: (str(v).strip() if v is not None else "") for i, v in enumerate(excel_row) if i < len(headers)}
            )
        return headers, rows
    finally:
        wb.close()


def resolve_present_in_db(result: Any, exists_fn: Callable[[int], bool]) -> tuple[str, str]:
    """
    Resolve the (mapped_title, present_in_db) tuple for a TitleMatchResult-shaped
    object.

    Prefers canonical_movie_id, falls back to suggested_movie_id. Returns the
    suggested title and "Yes" if the resolved id is > 0 and exists_fn(id) is
    True; otherwise returns ("", "No").
    """
    resolved_id: Optional[int] = getattr(result, "canonical_movie_id", None)
    if not resolved_id:
        resolved_id = getattr(result, "suggested_movie_id", None)

    if resolved_id and resolved_id > 0 and exists_fn(resolved_id):
        return getattr(result, "suggested_movie_title", "") or "", "Yes"

    return "", "No"


def build_output_xlsx(
    original_headers: list[str],
    rows: list[dict[str, Any]],
    results: list[dict[str, Any]],
) -> bytes:
    """
    Build an in-memory xlsx: one row per input row, columns = original_headers
    (in order) followed by mapped_title, confidence_score, reasoning,
    present_in_db.

    results[i] must be a dict with those 4 keys, corresponding to rows[i].
    """
    wb = openpyxl.Workbook()
    ws = wb.active

    out_headers = list(original_headers) + [
        "mapped_title",
        "confidence_score",
        "reasoning",
        "present_in_db",
    ]
    ws.append(out_headers)

    for row, result in zip(rows, results):
        out_row = [row.get(h, "") for h in original_headers]
        out_row += [
            result.get("mapped_title", ""),
            result.get("confidence_score", 0),
            result.get("reasoning", ""),
            result.get("present_in_db", "No"),
        ]
        ws.append(out_row)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def failed_row_result(message: str) -> dict[str, Any]:
    """Build the standard failed-row result dict."""
    return {
        "mapped_title": "",
        "confidence_score": 0,
        "reasoning": f"error: {message}",
        "present_in_db": "No",
    }
