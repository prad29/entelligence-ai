"""
Async batch worker for Phase 5.

Processes an uploaded .xlsx or .csv row-by-row, runs Layer 1 in-memory
detection, calls AWS Bedrock (Layer 2) for no-match rows when AI_TRIGGER_MODE
is enabled, writes an output .xlsx, and updates DetectionJob progress in the
database.
"""

import csv
import io
import json
import logging
import os
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timedelta
from typing import Any, Optional

import openpyxl
from openpyxl.styles import PatternFill
from sqlmodel import Session

from app.database import engine as db_engine
from app.models import DetectionJob, ReviewItem

logger = logging.getLogger(__name__)

_CHUNK_SIZE = 50


def _peek_headers(contents: bytes, ext: str) -> list[str]:
    """Return lowercased column headers from raw file bytes without writing to disk."""
    if ext == ".csv":
        import io as _io
        text = contents.decode("utf-8-sig", errors="replace")
        reader = csv.reader(_io.StringIO(text))
        raw = next(reader, [])
        return [h.strip().lower() for h in raw]
    wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
    ws = wb.active
    return [str(ws.cell(1, c).value or "").strip().lower() for c in range(1, (ws.max_column or 0) + 1)]


def _read_rows(upload_path: str) -> tuple[list[str], list[tuple]]:
    """
    Read amenity rows from either .xlsx or .csv.

    Returns (headers_lower, data_rows) where data_rows is a list of tuples
    with values in header order.
    """
    if upload_path.lower().endswith(".csv"):
        with open(upload_path, newline="", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            raw_headers = next(reader)
            headers = [h.strip().lower() for h in raw_headers]
            rows = [tuple(row) for row in reader]
        return headers, rows

    # xlsx
    wb = openpyxl.load_workbook(upload_path, data_only=True)
    ws = wb.active
    headers = [
        str(ws.cell(1, c).value or "").strip().lower()
        for c in range(1, ws.max_column + 1)
    ]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    return headers, rows


def run_batch_job(
    job_id: str,
    upload_path: str,
    include_diagnostics: bool,
    detection_engine,
) -> None:
    """
    Background task entry-point.  Called by FastAPI BackgroundTasks.

    Parameters
    ----------
    job_id:
        UUID string — primary key of the DetectionJob row.
    upload_path:
        Absolute path to the uploaded .xlsx saved by the endpoint.
    include_diagnostics:
        When True, extra columns (detected_keyword, match_source,
        match_track, confidence, ai_suggested_format, ai_reasoning)
        are appended to every output row.
    detection_engine:
        A ScreenFormatEngine instance loaded from app.state.engine.
    """
    from app.config import settings

    with Session(db_engine) as session:
        job = session.get(DetectionJob, job_id)
        if not job:
            logger.error("run_batch_job: job %s not found", job_id)
            return

        job.status = "processing"
        session.commit()

        try:
            _process_job(job_id, upload_path, include_diagnostics, detection_engine, session, settings)
        except Exception:
            job = session.get(DetectionJob, job_id)
            if job:
                job.status = "failed"
                session.commit()
            logger.exception("run_batch_job: job %s failed", job_id)
            raise


def _process_job(
    job_id: str,
    upload_path: str,
    include_diagnostics: bool,
    detection_engine,
    session: Session,
    settings,
) -> None:
    headers, rows = _read_rows(upload_path)
    amenities_idx = headers.index("amenities")
    circuit_idx = headers.index("circuit_name")

    # Update total to the actual row count (estimate may be off by ±1 for CSV)
    actual_total = len(rows)
    job = session.get(DetectionJob, job_id)
    if job and job.total != actual_total:
        job.total = actual_total
        session.commit()

    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    _AI_FILL = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")

    out_headers = ["circuit_name", "amenities", "screen_format"]
    if include_diagnostics:
        out_headers += [
            "detected_keyword",
            "match_source",
            "match_track",
            "confidence",
            "ai_suggested_format",
            "ai_reasoning",
        ]
    ws_out.append(out_headers)

    stats: dict[str, int] = {"matched": 0, "standard": 0, "ai_suggestions": 0, "no_match": 0}

    ai_enabled = settings.AI_TRIGGER_MODE not in ("off", "")
    pending_review: list[ReviewItem] = []

    if ai_enabled:
        from app.detection.bedrock_client import bedrock_client
        known_formats = detection_engine.get_all_formats()
    else:
        known_formats = []

    # --- Deduplication cache for Bedrock calls ---
    # Backed by Redis for cross-job persistence. Falls back to in-memory-only
    # if Redis is unavailable (non-fatal).
    dedup_cache: dict[tuple[str, str], Optional[Any]] = {}
    try:
        from app.cache import get_redis, bedrock_cache_key as _bck
        _redis_client = get_redis()
        _redis_client.ping()
        _redis_available = True
    except Exception:
        logger.warning("Redis unavailable — dedup cache will be in-memory only for this job")
        _redis_client = None
        _redis_available = False
    _ttl_seconds = settings.BEDROCK_CACHE_TTL_DAYS * 86400

    # =========================================================================
    # Pass 1: Run all rows through Layer 1 detection (fast, in-memory).
    # Separate rows into non-AI (immediately writable) and AI-pending.
    # =========================================================================

    # Storage for all row results, indexed by original row position (0-based).
    # Each entry: (amenity, circuit, result, needs_ai)
    row_data: list[tuple[str, str, Any, bool]] = []
    ai_pending: list[tuple[int, str, str, Any]] = []  # (row_idx_0based, amenity, circuit, result)

    for row_idx, row in enumerate(rows):
        if len(row) <= max(amenities_idx, circuit_idx):
            continue
        amenity = str(row[amenities_idx] or "").strip()
        circuit = str(row[circuit_idx] or "").strip()
        result = detection_engine.detect(amenity, circuit)

        if result.fired_ai:
            stats["no_match"] += 1
            stats["standard"] += 1
            row_data.append((amenity, circuit, result, True))
            ai_pending.append((row_idx, amenity, circuit, result))
        else:
            if result.screen_format != "Standard":
                stats["matched"] += 1
            else:
                stats["standard"] += 1
            row_data.append((amenity, circuit, result, False))

    # Write non-AI rows to the output sheet immediately and update progress.
    non_ai_count = 0
    for idx, (amenity, circuit, result, needs_ai) in enumerate(row_data):
        if not needs_ai:
            out_row: list = [circuit, amenity, result.screen_format]
            if include_diagnostics:
                out_row += [
                    result.detected_keyword or "",
                    result.match_source or "",
                    result.match_track or "",
                    result.confidence,
                    result.ai_suggested_format or "",
                    result.ai_reasoning or "",
                ]
            ws_out.append(out_row)
            # Track the output sheet row number for this data row
            row_data[idx] = (amenity, circuit, result, False)
            non_ai_count += 1

    # Update progress after pass 1
    job = session.get(DetectionJob, job_id)
    if job:
        job.processed = non_ai_count
    session.commit()

    # =========================================================================
    # Pass 2: Concurrent Bedrock calls for AI-pending rows.
    # Uses ThreadPoolExecutor with a semaphore to cap concurrency at 5.
    # Applies dedup_cache to skip redundant calls.
    # =========================================================================

    if ai_enabled and ai_pending:
        semaphore = threading.Semaphore(5)

        def _classify_with_semaphore(amenity: str, circuit: str):
            """Call Bedrock under semaphore; returns BedrockSuggestion or None."""
            with semaphore:
                return bedrock_client.classify_single(amenity, circuit, known_formats)

        # Build futures only for cache misses
        # future_map: future -> list of ai_pending indices that share this key
        future_map: dict[Any, list[int]] = {}
        # key_to_pending_indices: dedup key -> list of indices into ai_pending
        key_to_pending_indices: dict[tuple[str, str], list[int]] = {}

        for pending_idx, (row_idx_0, amenity, circuit, result) in enumerate(ai_pending):
            cache_key = (amenity.strip().lower(), circuit.strip().lower())
            if cache_key not in key_to_pending_indices:
                key_to_pending_indices[cache_key] = []
            key_to_pending_indices[cache_key].append(pending_idx)

        # Pre-populate dedup_cache from Redis for keys not already resolved
        if _redis_available:
            import json as _json
            from app.detection.types import BedrockSuggestion as _BS
            for cache_key in key_to_pending_indices:
                if cache_key in dedup_cache:
                    continue
                rkey = _bck(cache_key[0], cache_key[1])
                cached = _redis_client.get(rkey)
                if cached:
                    try:
                        dedup_cache[cache_key] = _BS(**_json.loads(cached))
                    except Exception:
                        pass  # malformed entry — let it re-call Bedrock

        with ThreadPoolExecutor(max_workers=5) as executor:
            futures = {}
            for cache_key, pending_indices in key_to_pending_indices.items():
                if cache_key in dedup_cache:
                    # Already resolved from Redis or prior in-memory hit
                    continue
                # Use the first pending item's raw amenity/circuit for the call
                first_pending_idx = pending_indices[0]
                amenity = ai_pending[first_pending_idx][1]
                circuit = ai_pending[first_pending_idx][2]
                future = executor.submit(_classify_with_semaphore, amenity, circuit)
                futures[future] = cache_key

            # Collect results as they complete
            completed_count = 0
            for future in as_completed(futures):
                cache_key = futures[future]
                try:
                    suggestion = future.result()
                except Exception:
                    logger.exception("Bedrock call failed for key %s", cache_key)
                    suggestion = None
                dedup_cache[cache_key] = suggestion

                # Persist to Redis so future jobs skip this Bedrock call
                if _redis_available and suggestion is not None:
                    try:
                        import json as _json
                        rkey = _bck(cache_key[0], cache_key[1])
                        _redis_client.setex(rkey, _ttl_seconds, _json.dumps(suggestion.__dict__))
                    except Exception:
                        logger.warning("Failed to write Bedrock result to Redis for key %s", cache_key)

                # Update progress incrementally
                completed_count += len(key_to_pending_indices[cache_key])
                job = session.get(DetectionJob, job_id)
                if job:
                    job.processed = non_ai_count + completed_count
                session.commit()

        # Apply cached results to all AI-pending rows
        for pending_idx, (row_idx_0, amenity, circuit, result) in enumerate(ai_pending):
            cache_key = (amenity.strip().lower(), circuit.strip().lower())
            suggestion = dedup_cache.get(cache_key)

            if suggestion:
                result.ai_suggested_format = suggestion.suggested_screen_format
                result.ai_reasoning = suggestion.reasoning
                stats["ai_suggestions"] += 1
                pending_review.append(ReviewItem(
                    type="ai_suggestion",
                    source_string=amenity,
                    circuit=circuit,
                    suggested_format=suggestion.suggested_screen_format,
                    confidence=suggestion.confidence,
                    reasoning=suggestion.reasoning,
                ))
            else:
                pending_review.append(ReviewItem(
                    type="ai_suggestion",
                    source_string=amenity,
                    circuit=circuit,
                    suggested_format="Standard",
                    confidence=0.0,
                    reasoning="Bedrock call failed — no match, Standard applied",
                ))

    elif not ai_enabled and ai_pending:
        # AI disabled but we have no-match rows — create review items
        for pending_idx, (row_idx_0, amenity, circuit, result) in enumerate(ai_pending):
            pending_review.append(ReviewItem(
                type="ai_suggestion",
                source_string=amenity,
                circuit=circuit,
                suggested_format="Standard",
                confidence=0.0,
                reasoning="No keyword match — AI trigger disabled",
            ))

    # =========================================================================
    # Write AI rows to output sheet in their original order.
    # =========================================================================

    for pending_idx, (row_idx_0, amenity, circuit, result) in enumerate(ai_pending):
        out_row = [circuit, amenity, result.screen_format]
        if include_diagnostics:
            out_row += [
                result.detected_keyword or "",
                result.match_source or "",
                result.match_track or "",
                result.confidence,
                result.ai_suggested_format or "",
                result.ai_reasoning or "",
            ]
        ws_out.append(out_row)

        # Highlight AI rows in light yellow
        row_num = ws_out.max_row
        for col in range(1, len(out_row) + 1):
            ws_out.cell(row=row_num, column=col).fill = _AI_FILL

    # Final: flush review items and mark job complete
    total_rows = len(row_data)
    job = session.get(DetectionJob, job_id)
    if job:
        job.processed = total_rows

    for ri in pending_review:
        session.add(ri)

    os.makedirs("/tmp/amenity_outputs", exist_ok=True)
    output_path = f"/tmp/amenity_outputs/{job_id}_output.xlsx"
    wb_out.save(output_path)

    if job:
        job.output_path = output_path
        job.status = "completed"
        job.stats = json.dumps(stats)
        job.ttl = datetime.utcnow() + timedelta(hours=settings.JOB_TTL_HOURS)

    session.commit()

    try:
        os.remove(upload_path)
    except OSError:
        pass

    logger.info(
        "run_batch_job: job %s completed — %d rows, stats=%s",
        job_id,
        total_rows,
        stats,
    )
