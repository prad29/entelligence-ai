import csv
import io
import json
import logging
import os
from datetime import datetime, timedelta
from typing import Any, Optional

import openpyxl
from sqlmodel import Session

from app.database import engine as db_engine
from app.models import IntlDetectionJob

logger = logging.getLogger(__name__)

_OUTPUT_DIR = "/tmp/intl_amenity_outputs"

# Update job.processed every N rows so the frontend progress bar moves,
# matching the domestic worker's cadence.
_PROGRESS_UPDATE_EVERY = 25


def _peek_headers(contents: bytes, ext: str) -> list[str]:
    """Copied verbatim from movie_batch_worker.py — duplicated rather than
    shared, consistent with how batch_worker.py and movie_batch_worker.py
    each carry their own copy."""
    if ext == ".csv":
        import io as _io
        text = contents.decode("utf-8-sig", errors="replace")
        reader = csv.reader(_io.StringIO(text))
        raw = next(reader, [])
        return [h.strip().lower() for h in raw]
    wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
    ws = wb.active
    return [str(ws.cell(1, c).value or "").strip().lower() for c in range(1, (ws.max_column or 0) + 1)]


def _read_rows(upload_path: str) -> tuple[list[str], list[tuple]]:
    """Copied verbatim from movie_batch_worker.py."""
    if upload_path.lower().endswith(".csv"):
        with open(upload_path, newline="", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            raw_headers = next(reader)
            headers = [h.strip().lower() for h in raw_headers]
            rows = [tuple(row) for row in reader]
        return headers, rows

    wb = openpyxl.load_workbook(upload_path, data_only=True)
    ws = wb.active
    headers = [
        str(ws.cell(1, c).value or "").strip().lower()
        for c in range(1, ws.max_column + 1)
    ]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    return headers, rows


_SENTINEL_VALUES = {"undefined", "null", "none", "n/a", "na"}


def run_intl_batch_job(
    job_id: str,
    upload_path: str,
    include_diagnostics: bool,
    detection_engine,
    audit_mode: bool = False,
) -> None:
    """Rule-engine-only batch pass for international amenity detection.

    No ThreadPoolExecutor, no threading.Semaphore, no PatternFill AI
    highlighting, no Bedrock — nothing is AI-classified in this build, so
    that machinery from movie_batch_worker.py is intentionally not ported.
    """
    from app.config import settings

    with Session(db_engine) as session:
        job = session.get(IntlDetectionJob, job_id)
        if not job:
            logger.error("run_intl_batch_job: job %s not found", job_id)
            return

        job.status = "processing"
        session.commit()

        try:
            _process_job(job_id, upload_path, include_diagnostics, detection_engine, session, settings, audit_mode)
        except Exception:
            job = session.get(IntlDetectionJob, job_id)
            if job:
                job.status = "failed"
                job.stats = json.dumps({"error": "processing failed"})
                session.commit()
            logger.exception("run_intl_batch_job: job %s failed", job_id)
        finally:
            if os.path.exists(upload_path):
                try:
                    os.remove(upload_path)
                except OSError:
                    pass


def _process_job(
    job_id: str,
    upload_path: str,
    include_diagnostics: bool,
    detection_engine,
    session: Session,
    settings,
    audit_mode: bool = False,
) -> None:
    headers, rows = _read_rows(upload_path)

    if "amenities_string" in headers:
        amenities_idx = headers.index("amenities_string")
    elif "amenities" in headers:
        amenities_idx = headers.index("amenities")
    else:
        job = session.get(IntlDetectionJob, job_id)
        if job:
            job.status = "failed"
            job.stats = json.dumps({"error": "missing an amenities or amenities_string column"})
            session.commit()
        logger.error("_process_job: job %s missing amenities column", job_id)
        return

    user_format_idx: Optional[int] = None
    if audit_mode:
        if "screen_format" not in headers:
            job = session.get(IntlDetectionJob, job_id)
            if job:
                job.status = "failed"
                job.stats = json.dumps({"error": "audit_mode requires column: screen_format"})
                session.commit()
            logger.error("_process_job: job %s missing screen_format column", job_id)
            return
        user_format_idx = headers.index("screen_format")

    actual_total = len(rows)
    job = session.get(IntlDetectionJob, job_id)
    if job and job.total != actual_total:
        job.total = actual_total
        session.commit()

    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active

    out_headers = list(headers) + [
        "screen_format",
        "match_track",
        "confidence",
        "matched_keyword",
        "priority_tier",
        "match_source",
    ]
    if include_diagnostics:
        out_headers += ["diagnostics"]
    if audit_mode:
        out_headers += ["match_status"]
    ws_out.append(out_headers)

    stats: dict[str, Any] = {"matched": 0, "no_match": 0}
    if audit_mode:
        stats["mismatch_count"] = 0

    for row_idx, row in enumerate(rows):
        amenity = str(row[amenities_idx] if len(row) > amenities_idx else "").strip()
        if amenity.lower() in _SENTINEL_VALUES:
            amenity = ""

        result = detection_engine.detect(amenity)

        if result.match_source == "Keyword Match":
            stats["matched"] += 1
        else:
            stats["no_match"] += 1

        out_row: list = list(row) + [
            result.screen_format,
            result.match_track,
            result.confidence,
            result.matched_keyword or "",
            result.priority_tier,
            result.match_source or "",
        ]
        if include_diagnostics:
            out_row += [json.dumps(result.diagnostics) if result.diagnostics else ""]
        if audit_mode:
            user_format = str(row[user_format_idx] if user_format_idx is not None and len(row) > user_format_idx else "").strip()
            match_status = "MATCH" if user_format.strip().lower() == result.screen_format.strip().lower() else "MISMATCH"
            if match_status == "MISMATCH":
                stats["mismatch_count"] += 1
            out_row += [match_status]

        ws_out.append(out_row)

        if (row_idx + 1) % _PROGRESS_UPDATE_EVERY == 0:
            job = session.get(IntlDetectionJob, job_id)
            if job:
                job.processed = row_idx + 1
            session.commit()

    total_rows = len(rows)
    job = session.get(IntlDetectionJob, job_id)
    if job:
        job.processed = total_rows

    os.makedirs(_OUTPUT_DIR, exist_ok=True)
    output_path = f"{_OUTPUT_DIR}/{job_id}_output.xlsx"
    wb_out.save(output_path)

    if job:
        job.output_path = output_path
        job.status = "completed"
        job.stats = json.dumps(stats)
        job.ttl = datetime.utcnow() + timedelta(hours=settings.JOB_TTL_HOURS)

    session.commit()

    logger.info(
        "run_intl_batch_job: job %s completed — %d rows, stats=%s",
        job_id,
        total_rows,
        stats,
    )
