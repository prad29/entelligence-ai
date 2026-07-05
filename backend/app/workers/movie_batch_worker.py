import csv
import io
import json
import logging
import os
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timedelta
from typing import Any, Optional

import openpyxl
from openpyxl.styles import PatternFill
from sqlmodel import Session

from app.database import engine as db_engine
from app.models import MovieFormatJob, MovieFormatReviewItem

logger = logging.getLogger(__name__)

_KNOWN_FORMATS = ["70MM", "35MM", "3D", "2D"]


def _peek_headers(contents: bytes, ext: str) -> list[str]:
    if ext == ".csv":
        import io as _io
        text = contents.decode("utf-8-sig", errors="replace")
        reader = csv.reader(_io.StringIO(text))
        raw = next(reader, [])
        return [h.strip().lower() for h in raw]
    wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
    ws = wb.active
    return [str(ws.cell(1, c).value or "").strip().lower() for c in range(1, (ws.max_column or 0) + 1)]


def _read_rows(upload_path: str) -> tuple[list[str], list[tuple]]:
    if upload_path.lower().endswith(".csv"):
        with open(upload_path, newline="", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            raw_headers = next(reader)
            headers = [h.strip().lower() for h in raw_headers]
            rows = [tuple(row) for row in reader]
        return headers, rows

    wb = openpyxl.load_workbook(upload_path, data_only=True)
    ws = wb.active
    headers = [
        str(ws.cell(1, c).value or "").strip().lower()
        for c in range(1, ws.max_column + 1)
    ]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    return headers, rows


def run_movie_batch_job(
    job_id: str,
    upload_path: str,
    include_diagnostics: bool,
    detection_engine,
    batch_ai_mode: str = "skip",
) -> None:
    from app.config import settings

    with Session(db_engine) as session:
        job = session.get(MovieFormatJob, job_id)
        if not job:
            logger.error("run_movie_batch_job: job %s not found", job_id)
            return

        job.status = "processing"
        session.commit()

        try:
            _process_job(job_id, upload_path, include_diagnostics, detection_engine, session, settings, batch_ai_mode)
        except Exception:
            job = session.get(MovieFormatJob, job_id)
            if job:
                job.status = "failed"
                session.commit()
            logger.exception("run_movie_batch_job: job %s failed", job_id)
            raise


def _process_job(
    job_id: str,
    upload_path: str,
    include_diagnostics: bool,
    detection_engine,
    session: Session,
    settings,
    batch_ai_mode: str = "skip",
) -> None:
    headers, rows = _read_rows(upload_path)
    amenities_idx = headers.index("amenities")

    actual_total = len(rows)
    job = session.get(MovieFormatJob, job_id)
    if job and job.total != actual_total:
        job.total = actual_total
        session.commit()

    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    _AI_FILL = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")

    out_headers = ["amenities", "movie_format"]
    if include_diagnostics:
        out_headers += [
            "detected_keyword",
            "match_source",
            "match_track",
            "confidence",
            "ai_suggested_format",
            "ai_reasoning",
        ]
    ws_out.append(out_headers)

    stats: dict[str, int] = {"matched": 0, "standard": 0, "ai_suggestions": 0, "no_match": 0}

    ai_enabled = settings.AI_TRIGGER_MODE not in ("off", "")
    pending_review: list[MovieFormatReviewItem] = []

    if ai_enabled:
        from app.detection.bedrock_client import bedrock_client

    dedup_cache: dict[str, Optional[Any]] = {}
    try:
        from app.cache import get_redis, movie_format_cache_key as _mck
        _redis_client = get_redis()
        _redis_client.ping()
        _redis_available = True
    except Exception:
        logger.warning("Redis unavailable — dedup cache will be in-memory only for this job")
        _redis_client = None
        _redis_available = False
    _ttl_seconds = settings.BEDROCK_CACHE_TTL_DAYS * 86400

    # Pass 1: Layer 1 detection on all rows
    row_data: list[tuple[str, Any, bool]] = []
    ai_pending: list[tuple[int, str, Any]] = []  # (row_idx_0based, amenity, result)

    for row_idx, row in enumerate(rows):
        amenity = str(row[amenities_idx] if len(row) > amenities_idx else "").strip()
        result = detection_engine.detect(amenity)

        if result.fired_ai:
            stats["no_match"] += 1
            stats["standard"] += 1
            row_data.append((amenity, result, True))
            ai_pending.append((row_idx, amenity, result))
        else:
            if result.movie_format != "2D":
                stats["matched"] += 1
            else:
                stats["standard"] += 1
            row_data.append((amenity, result, False))

    # Write non-AI rows immediately
    non_ai_count = 0
    for idx, (amenity, result, needs_ai) in enumerate(row_data):
        if not needs_ai:
            out_row: list = [amenity, result.movie_format]
            if include_diagnostics:
                out_row += [
                    result.detected_keyword or "",
                    result.match_source or "",
                    result.match_track or "",
                    result.confidence,
                    result.ai_suggested_format or "",
                    result.ai_reasoning or "",
                ]
            ws_out.append(out_row)
            non_ai_count += 1

    job = session.get(MovieFormatJob, job_id)
    if job:
        job.processed = non_ai_count
    session.commit()

    # Pass 2: Concurrent Bedrock calls for AI-pending rows
    effective_ai = ai_enabled and batch_ai_mode == "full"
    sample_mode = ai_enabled and batch_ai_mode == "sample"
    sample_limit = getattr(settings, "BATCH_AI_SAMPLE_LIMIT", 50)

    if effective_ai and ai_pending:
        _concurrency = getattr(settings, "BEDROCK_MAX_CONCURRENCY", 20)
        semaphore = threading.Semaphore(_concurrency)

        def _classify_with_semaphore(amenity: str):
            with semaphore:
                return bedrock_client.classify_single(amenity, "", _KNOWN_FORMATS)

        key_to_pending_indices: dict[str, list[int]] = {}
        for pending_idx, (row_idx_0, amenity, result) in enumerate(ai_pending):
            cache_key = amenity.strip().lower()
            if cache_key not in key_to_pending_indices:
                key_to_pending_indices[cache_key] = []
            key_to_pending_indices[cache_key].append(pending_idx)

        if _redis_available:
            import json as _json
            from app.movie_detection.types import MovieFormatBedrockSuggestion as _BS
            uncached_keys = [k for k in key_to_pending_indices if k not in dedup_cache]
            if uncached_keys:
                rkeys = [_mck(k) for k in uncached_keys]
                values = _redis_client.mget(rkeys)
                for cache_key, raw in zip(uncached_keys, values):
                    if raw:
                        try:
                            dedup_cache[cache_key] = _BS(**_json.loads(raw))
                        except Exception:
                            pass

        with ThreadPoolExecutor(max_workers=_concurrency) as executor:
            futures = {}
            for cache_key, pending_indices in key_to_pending_indices.items():
                if cache_key in dedup_cache:
                    continue
                first_pending_idx = pending_indices[0]
                amenity = ai_pending[first_pending_idx][1]
                future = executor.submit(_classify_with_semaphore, amenity)
                futures[future] = cache_key

            completed_count = 0
            for future in as_completed(futures):
                cache_key = futures[future]
                try:
                    suggestion = future.result()
                except Exception:
                    logger.exception("Bedrock call failed for key %s", cache_key)
                    suggestion = None
                dedup_cache[cache_key] = suggestion

                if _redis_available and suggestion is not None:
                    try:
                        import json as _json
                        rkey = _mck(cache_key)
                        _redis_client.setex(rkey, _ttl_seconds, _json.dumps(suggestion.__dict__))
                    except Exception:
                        logger.warning("Failed to write Bedrock result to Redis for key %s", cache_key)

                completed_count += len(key_to_pending_indices[cache_key])
                job = session.get(MovieFormatJob, job_id)
                if job:
                    job.processed = non_ai_count + completed_count
                session.commit()

        for pending_idx, (row_idx_0, amenity, result) in enumerate(ai_pending):
            cache_key = amenity.strip().lower()
            suggestion = dedup_cache.get(cache_key)

            if suggestion:
                result.ai_suggested_format = suggestion.suggested_screen_format
                result.ai_reasoning = suggestion.reasoning
                stats["ai_suggestions"] += 1
                pending_review.append(MovieFormatReviewItem(
                    type="ai_suggestion",
                    source_string=amenity,
                    suggested_format=suggestion.suggested_screen_format,
                    confidence=suggestion.confidence,
                    reasoning=suggestion.reasoning,
                ))
            else:
                pending_review.append(MovieFormatReviewItem(
                    type="ai_suggestion",
                    source_string=amenity,
                    suggested_format="2D",
                    confidence=0.0,
                    reasoning="Bedrock call failed — no match, 2D applied",
                ))

    elif sample_mode and ai_pending:
        _concurrency_s = getattr(settings, "BEDROCK_MAX_CONCURRENCY", 20)
        semaphore_s = threading.Semaphore(_concurrency_s)

        def _classify_sample(amenity: str):
            with semaphore_s:
                return bedrock_client.classify_single(amenity, "", _KNOWN_FORMATS)

        key_to_pending_indices: dict[str, list[int]] = {}
        for pending_idx, (row_idx_0, amenity, result) in enumerate(ai_pending):
            cache_key = amenity.strip().lower()
            if cache_key not in key_to_pending_indices:
                key_to_pending_indices[cache_key] = []
            key_to_pending_indices[cache_key].append(pending_idx)

        sampled_keys = dict(list(key_to_pending_indices.items())[:sample_limit])
        with ThreadPoolExecutor(max_workers=_concurrency_s) as executor_s:
            futures_s = {
                executor_s.submit(_classify_sample, ai_pending[idxs[0]][1]): ck
                for ck, idxs in sampled_keys.items()
                if ck not in dedup_cache
            }
            for future_s in as_completed(futures_s):
                ck = futures_s[future_s]
                try:
                    dedup_cache[ck] = future_s.result()
                except Exception:
                    dedup_cache[ck] = None

        for pending_idx, (row_idx_0, amenity, result) in enumerate(ai_pending):
            cache_key = amenity.strip().lower()
            suggestion = dedup_cache.get(cache_key)
            if suggestion:
                result.ai_suggested_format = suggestion.suggested_screen_format
                result.ai_reasoning = suggestion.reasoning
                stats["ai_suggestions"] += 1
                pending_review.append(MovieFormatReviewItem(
                    type="ai_suggestion",
                    source_string=amenity,
                    suggested_format=suggestion.suggested_screen_format,
                    confidence=suggestion.confidence,
                    reasoning=suggestion.reasoning,
                ))
            else:
                pending_review.append(MovieFormatReviewItem(
                    type="ai_suggestion",
                    source_string=amenity,
                    suggested_format="2D",
                    confidence=0.0,
                    reasoning="No keyword match — outside AI sample limit",
                ))

    elif (not effective_ai) and ai_pending:
        for pending_idx, (row_idx_0, amenity, result) in enumerate(ai_pending):
            pending_review.append(MovieFormatReviewItem(
                type="ai_suggestion",
                source_string=amenity,
                suggested_format="2D",
                confidence=0.0,
                reasoning="No keyword match — AI trigger disabled",
            ))

    # Write AI rows to output
    for pending_idx, (row_idx_0, amenity, result) in enumerate(ai_pending):
        out_row = [amenity, result.movie_format]
        if include_diagnostics:
            out_row += [
                result.detected_keyword or "",
                result.match_source or "",
                result.match_track or "",
                result.confidence,
                result.ai_suggested_format or "",
                result.ai_reasoning or "",
            ]
        ws_out.append(out_row)

        row_num = ws_out.max_row
        for col in range(1, len(out_row) + 1):
            ws_out.cell(row=row_num, column=col).fill = _AI_FILL

    total_rows = len(row_data)
    job = session.get(MovieFormatJob, job_id)
    if job:
        job.processed = total_rows

    for ri in pending_review:
        session.add(ri)

    os.makedirs("/tmp/movie_outputs", exist_ok=True)
    output_path = f"/tmp/movie_outputs/{job_id}_output.xlsx"
    wb_out.save(output_path)

    if job:
        job.output_path = output_path
        job.status = "completed"
        job.stats = json.dumps(stats)
        job.ttl = datetime.utcnow() + timedelta(hours=settings.JOB_TTL_HOURS)

    session.commit()

    try:
        os.remove(upload_path)
    except OSError:
        pass

    if ai_pending:
        from collections import Counter
        unmatched_counter = Counter(amenity.strip().lower() for (_, amenity, _) in ai_pending)
        top_unmatched = unmatched_counter.most_common(10)
        logger.info("movie_batch_top_unmatched job=%s: %s", job_id, top_unmatched)
        stats["top_unmatched"] = [{"amenity": a, "count": c} for a, c in top_unmatched]

    logger.info(
        "run_movie_batch_job: job %s completed — %d rows, stats=%s",
        job_id,
        total_rows,
        stats,
    )
