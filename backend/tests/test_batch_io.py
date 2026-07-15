"""Unit tests for app.title_matching.batch_io — pure helpers, no DB/network."""

import io
from dataclasses import dataclass
from typing import Optional

import openpyxl
import pytest

from app.title_matching.batch_io import (
    REQUIRED_COLUMNS,
    build_output_xlsx,
    failed_row_result,
    parse_upload,
    peek_headers,
    resolve_present_in_db,
)


@dataclass
class _FakeResult:
    suggested_movie_id: int
    suggested_movie_title: str
    canonical_movie_id: Optional[int] = None


def _csv_bytes(header_line: str, *data_lines: str) -> bytes:
    text = "\n".join([header_line, *data_lines]) + "\n"
    return text.encode("utf-8-sig")


def _xlsx_bytes(headers: list[str], rows: list[list]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class TestParseUploadCSV:
    def test_parses_csv_with_required_columns_present(self):
        contents = _csv_bytes(
            "movie_title,show_date,ticketing_url,theater",
            "Dune,2024-01-01,https://example.com/dune,Cineplex",
        )

        headers, rows = parse_upload(contents, ".csv")

        assert headers == ["movie_title", "show_date", "ticketing_url", "theater"]
        assert len(rows) == 1
        assert rows[0]["movie_title"] == "Dune"
        assert rows[0]["show_date"] == "2024-01-01"
        assert rows[0]["ticketing_url"] == "https://example.com/dune"

    def test_missing_required_column_raises_value_error(self):
        contents = _csv_bytes(
            "movie_title,ticketing_url",
            "Dune,https://example.com/dune",
        )

        with pytest.raises(ValueError, match="show_date"):
            parse_upload(contents, ".csv")

    def test_required_columns_matched_case_insensitively(self):
        contents = _csv_bytes(
            "Movie_Title,Show_Date,Ticketing_URL",
            "Dune,2024-01-01,https://example.com/dune",
        )

        headers, rows = parse_upload(contents, ".csv")

        assert headers == ["Movie_Title", "Show_Date", "Ticketing_URL"]
        assert len(rows) == 1

    def test_title_column_accepted_as_movie_title_alias(self):
        contents = _csv_bytes(
            "title,master_title,theater_name,show_date,ticketing_url",
            "Oh Sukumari,Oh..! Sukumari,AMC Concord Mills 24,2026-07-21,https://example.com/oh",
        )

        headers, rows = parse_upload(contents, ".csv")

        assert headers == ["title", "master_title", "theater_name", "show_date", "ticketing_url"]
        assert len(rows) == 1
        assert rows[0]["title"] == "Oh Sukumari"


class TestParseUploadXLSX:
    def test_parses_xlsx_with_required_columns_present(self):
        contents = _xlsx_bytes(
            ["movie_title", "show_date", "ticketing_url"],
            [["Dune", "2024-01-01", "https://example.com/dune"]],
        )

        headers, rows = parse_upload(contents, ".xlsx")

        assert headers == ["movie_title", "show_date", "ticketing_url"]
        assert len(rows) == 1
        assert rows[0]["movie_title"] == "Dune"
        assert rows[0]["ticketing_url"] == "https://example.com/dune"

    def test_missing_required_column_raises_value_error(self):
        contents = _xlsx_bytes(
            ["movie_title", "theater"],
            [["Dune", "Cineplex"]],
        )

        with pytest.raises(ValueError, match="show_date"):
            parse_upload(contents, ".xlsx")

    def test_trailing_blank_rows_from_stale_dimension_metadata_are_skipped(self):
        # Simulates a workbook whose stored <dimension> range is larger than
        # its actual populated rows (e.g. after rows were deleted in Excel) —
        # openpyxl's max_row/iter_rows still walks the stale range and yields
        # fully-blank rows that must not become phantom entries.
        contents = _xlsx_bytes(
            ["movie_title", "show_date", "ticketing_url"],
            [
                ["Dune", "2024-01-01", "https://example.com/dune"],
                [None, None, None],
                [None, None, None],
            ],
        )

        headers, rows = parse_upload(contents, ".xlsx")

        assert len(rows) == 1
        assert rows[0]["movie_title"] == "Dune"


class TestPeekHeaders:
    def test_peek_headers_csv(self):
        contents = _csv_bytes("movie_title,show_date,ticketing_url")
        assert peek_headers(contents, ".csv") == ["movie_title", "show_date", "ticketing_url"]

    def test_peek_headers_xlsx(self):
        contents = _xlsx_bytes(["movie_title", "show_date", "ticketing_url"], [])
        assert peek_headers(contents, ".xlsx") == ["movie_title", "show_date", "ticketing_url"]


class TestResolvePresentInDb:
    def test_id_present_and_exists_returns_title_and_yes(self):
        result = _FakeResult(suggested_movie_id=42, suggested_movie_title="Dune", canonical_movie_id=42)

        mapped_title, present_flag = resolve_present_in_db(result, exists_fn=lambda mid: True)

        assert mapped_title == "Dune"
        assert present_flag == "Yes"

    def test_id_zero_returns_empty_and_no(self):
        result = _FakeResult(suggested_movie_id=0, suggested_movie_title="", canonical_movie_id=0)

        mapped_title, present_flag = resolve_present_in_db(result, exists_fn=lambda mid: True)

        assert mapped_title == ""
        assert present_flag == "No"

    def test_id_positive_but_not_exists_returns_empty_and_no(self):
        result = _FakeResult(suggested_movie_id=99, suggested_movie_title="Ghost", canonical_movie_id=99)

        mapped_title, present_flag = resolve_present_in_db(result, exists_fn=lambda mid: False)

        assert mapped_title == ""
        assert present_flag == "No"

    def test_prefers_canonical_id_over_suggested_id(self):
        seen_ids = []

        def exists_fn(mid):
            seen_ids.append(mid)
            return True

        result = _FakeResult(suggested_movie_id=5, suggested_movie_title="Dune Part Two", canonical_movie_id=7)

        mapped_title, present_flag = resolve_present_in_db(result, exists_fn=exists_fn)

        assert seen_ids == [7]
        assert mapped_title == "Dune Part Two"
        assert present_flag == "Yes"

    def test_falls_back_to_suggested_id_when_canonical_missing(self):
        seen_ids = []

        def exists_fn(mid):
            seen_ids.append(mid)
            return True

        result = _FakeResult(suggested_movie_id=5, suggested_movie_title="Dune", canonical_movie_id=0)

        mapped_title, present_flag = resolve_present_in_db(result, exists_fn=exists_fn)

        assert seen_ids == [5]
        assert mapped_title == "Dune"
        assert present_flag == "Yes"


class TestFailedRowResult:
    def test_shape_of_failed_row_result(self):
        result = failed_row_result("timeout after 90s")

        assert result == {
            "mapped_title": "",
            "confidence_score": 0,
            "reasoning": "error: timeout after 90s",
            "present_in_db": "No",
        }


class TestBuildOutputXlsx:
    def test_round_trips_with_original_plus_four_appended_columns(self):
        original_headers = ["movie_title", "show_date", "ticketing_url"]
        rows = [
            {"movie_title": "Dune", "show_date": "2024-01-01", "ticketing_url": "https://example.com/dune"},
            {"movie_title": "Unknown Movie", "show_date": "2024-02-02", "ticketing_url": "https://example.com/unknown"},
        ]
        results = [
            {
                "mapped_title": "Dune",
                "confidence_score": 0.95,
                "reasoning": "exact match",
                "present_in_db": "Yes",
            },
            failed_row_result("agentic sandbox timeout"),
        ]

        xlsx_bytes = build_output_xlsx(original_headers, rows, results)

        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
        ws = wb.active
        rows_out = list(ws.iter_rows(values_only=True))

        expected_headers = (
            "movie_title",
            "show_date",
            "ticketing_url",
            "mapped_title",
            "confidence_score",
            "reasoning",
            "present_in_db",
        )
        assert rows_out[0] == expected_headers

        assert rows_out[1] == (
            "Dune",
            "2024-01-01",
            "https://example.com/dune",
            "Dune",
            0.95,
            "exact match",
            "Yes",
        )

        # Note: openpyxl round-trips an empty string cell as None on reload —
        # this is standard openpyxl behavior, not a defect in build_output_xlsx.
        assert rows_out[2] == (
            "Unknown Movie",
            "2024-02-02",
            "https://example.com/unknown",
            None,
            0,
            "error: agentic sandbox timeout",
            "No",
        )

    def test_empty_rows_produces_only_header(self):
        xlsx_bytes = build_output_xlsx(["movie_title"], [], [])

        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
        ws = wb.active
        rows_out = list(ws.iter_rows(values_only=True))

        assert rows_out == [("movie_title", "mapped_title", "confidence_score", "reasoning", "present_in_db")]


def test_required_columns_constant():
    assert REQUIRED_COLUMNS == ("movie_title", "show_date", "ticketing_url")
