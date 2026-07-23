"""
API tests for the batch title-matching endpoints:

    POST /api/v1/movie-title-match/batch
    GET  /api/v1/movie-title-match/batch/{job_id}
    GET  /api/v1/movie-title-match/batch/{job_id}/download

dispatch_batch is monkeypatched to a no-op stub so these tests never touch
Celery, Redis, or the claude-sandbox subprocess — pure HTTP + DB behavior.
batch_storage (S3) is monkeypatched to an in-memory dict so these tests never
touch real S3 either.

Uses an in-memory SQLite DB (mirrors tests/test_api_integration.py's pattern)
so the suite runs without PostgreSQL/Docker.
"""

from __future__ import annotations

import io
from datetime import datetime, timedelta

import pytest
from fastapi.testclient import TestClient
from sqlmodel import Session, SQLModel, create_engine
from sqlalchemy.pool import StaticPool

import app.database as _db_module
from app.config import settings
from app.database import get_session
from app.main import app
from app.models import MovieTitleBatchJob, MovieTitleIntlBatchJob

# ── shared in-memory sqlite engine ───────────────────────────────────────────
_sqlite_engine = create_engine(
    "sqlite:///:memory:",
    connect_args={"check_same_thread": False},
    poolclass=StaticPool,
)
_db_module.engine = _sqlite_engine


def _override_get_session():
    with Session(_sqlite_engine) as session:
        yield session


app.dependency_overrides[get_session] = _override_get_session
SQLModel.metadata.create_all(_sqlite_engine)

client = TestClient(app, raise_server_exceptions=False)


@pytest.fixture(autouse=True)
def _clean_tables():
    """Wipe MovieTitleBatchJob/MovieTitleIntlBatchJob rows between tests for isolation."""
    with Session(_sqlite_engine) as session:
        for row in session.exec(
            __import__("sqlmodel").select(MovieTitleBatchJob)
        ).all():
            session.delete(row)
        for row in session.exec(
            __import__("sqlmodel").select(MovieTitleIntlBatchJob)
        ).all():
            session.delete(row)
        session.commit()
    yield


@pytest.fixture(autouse=True)
def _agentic_enabled(monkeypatch):
    """Batch matching requires Mode B enabled — default ON for most tests."""
    monkeypatch.setattr(settings, "AGENTIC_TITLE_MATCH_ENABLED", True)
    yield


@pytest.fixture(autouse=True)
def _stub_dispatch_batch(monkeypatch):
    """Never touch real Celery/Redis/sandbox — dispatch_batch_task.delay is a no-op.

    The router enqueues dispatch_batch_task rather than calling dispatch_batch
    inline (see movie_title_match.py), so stubbing .delay is what actually
    intercepts the call in these HTTP+DB-only tests.
    """
    calls = []

    def _fake_delay(job_id):
        calls.append(job_id)

    monkeypatch.setattr(
        "app.tasks.agentic_match_task.dispatch_batch_task.delay", _fake_delay
    )
    return calls


@pytest.fixture(autouse=True)
def _stub_dispatch_intl_batch(monkeypatch):
    """Never touch real Celery/Redis/sandbox for the international path either."""
    calls = []

    def _fake_delay(job_id):
        calls.append(job_id)

    monkeypatch.setattr(
        "app.tasks.agentic_intl_match_task.dispatch_intl_batch_task.delay", _fake_delay
    )
    return calls


@pytest.fixture(autouse=True)
def _stub_batch_storage(monkeypatch):
    """Never touch real S3 — batch_storage backed by an in-memory dict."""
    import app.title_matching.batch_storage as storage

    store: dict[str, bytes] = {}

    monkeypatch.setattr(storage, "put_bytes", lambda key, data: store.__setitem__(key, data))
    monkeypatch.setattr(storage, "get_bytes", lambda key: store[key])
    monkeypatch.setattr(storage, "delete", lambda key: store.pop(key, None))
    monkeypatch.setattr(storage, "exists", lambda key: key in store)
    return store


def _csv_bytes(*lines: str) -> bytes:
    return ("\n".join(lines) + "\n").encode("utf-8-sig")


def _valid_csv() -> bytes:
    return _csv_bytes(
        "movie_title,show_date,ticketing_url",
        "Dune,2024-01-01,https://example.com/dune",
        "Oppenheimer,2023-07-21,https://example.com/opp",
    )


def _valid_intl_csv() -> bytes:
    return _csv_bytes(
        "movie_title,show_date,ticketing_url,country",
        "Blue Beetle,2023-08-16,https://example.com/blue-beetle,France",
    )


def _get_job(job_id: str) -> MovieTitleBatchJob:
    with Session(_sqlite_engine) as session:
        return session.get(MovieTitleBatchJob, job_id)


def _get_intl_job(job_id: str) -> MovieTitleIntlBatchJob:
    with Session(_sqlite_engine) as session:
        return session.get(MovieTitleIntlBatchJob, job_id)


# ─────────────────────────────────────────────────────────────────────────────
# POST /batch
# ─────────────────────────────────────────────────────────────────────────────

def test_valid_csv_upload_returns_200_with_job_id():
    resp = client.post(
        "/api/v1/movie-title-match/batch",
        files={"file": ("titles.csv", _valid_csv(), "text/csv")},
        data={"use_poster_vision": "false"},
    )
    assert resp.status_code == 200
    body = resp.json()
    assert "job_id" in body and body["job_id"]

    job = _get_job(body["job_id"])
    assert job is not None
    assert job.total == 2
    assert job.use_poster_vision is False


def test_upload_missing_required_column_returns_400():
    bad_csv = _csv_bytes(
        "movie_title,ticketing_url",
        "Dune,https://example.com/dune",
    )
    resp = client.post(
        "/api/v1/movie-title-match/batch",
        files={"file": ("titles.csv", bad_csv, "text/csv")},
    )
    assert resp.status_code == 400
    assert "show_date" in resp.json()["detail"]


def test_upload_when_agentic_disabled_returns_400(monkeypatch):
    monkeypatch.setattr(settings, "AGENTIC_TITLE_MATCH_ENABLED", False)
    resp = client.post(
        "/api/v1/movie-title-match/batch",
        files={"file": ("titles.csv", _valid_csv(), "text/csv")},
    )
    assert resp.status_code == 400
    assert "Mode B" in resp.json()["detail"] or "agentic" in resp.json()["detail"].lower()


def test_upload_bad_extension_returns_400():
    resp = client.post(
        "/api/v1/movie-title-match/batch",
        files={"file": ("titles.txt", b"movie_title,show_date,ticketing_url\n", "text/plain")},
    )
    assert resp.status_code == 400


# ─────────────────────────────────────────────────────────────────────────────
# GET /batch/{job_id}
# ─────────────────────────────────────────────────────────────────────────────

def test_status_on_fresh_job_has_zero_progress_not_nan():
    with Session(_sqlite_engine) as session:
        job = MovieTitleBatchJob(id="job-fresh", status="queued", total=0)
        session.add(job)
        session.commit()

    resp = client.get("/api/v1/movie-title-match/batch/job-fresh")
    assert resp.status_code == 200
    body = resp.json()
    assert body["processed"] == 0
    assert body["progress"] == 0
    assert body["progress"] == body["progress"]  # NaN != NaN; this proves not-NaN


def test_status_progress_computed_when_total_positive():
    with Session(_sqlite_engine) as session:
        job = MovieTitleBatchJob(id="job-progress", status="processing", total=4, processed=1)
        session.add(job)
        session.commit()

    resp = client.get("/api/v1/movie-title-match/batch/job-progress")
    assert resp.status_code == 200
    assert resp.json()["progress"] == 0.25


def test_status_unknown_job_returns_404():
    resp = client.get("/api/v1/movie-title-match/batch/does-not-exist")
    assert resp.status_code == 404


# ─────────────────────────────────────────────────────────────────────────────
# GET /batch/{job_id}/download
# ─────────────────────────────────────────────────────────────────────────────

def test_download_before_completion_returns_400():
    with Session(_sqlite_engine) as session:
        job = MovieTitleBatchJob(id="job-notdone", status="processing", total=2)
        session.add(job)
        session.commit()

    resp = client.get("/api/v1/movie-title-match/batch/job-notdone/download")
    assert resp.status_code == 400


def test_download_after_completion_returns_200_xlsx(_stub_batch_storage):
    import io as _io

    import openpyxl

    wb = openpyxl.Workbook()
    wb.active.append(["movie_title", "mapped_title"])
    buf = _io.BytesIO()
    wb.save(buf)
    output_key = "batch-outputs/job-done_output.xlsx"
    _stub_batch_storage[output_key] = buf.getvalue()

    with Session(_sqlite_engine) as session:
        job = MovieTitleBatchJob(
            id="job-done",
            status="completed",
            total=1,
            processed=1,
            matched=1,
            output_path=output_key,
            ttl=datetime.utcnow() + timedelta(hours=1),
        )
        session.add(job)
        session.commit()

    resp = client.get("/api/v1/movie-title-match/batch/job-done/download")
    assert resp.status_code == 200
    assert resp.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    # status endpoint should now surface output_url
    status_resp = client.get("/api/v1/movie-title-match/batch/job-done")
    assert status_resp.json()["output_url"] == (
        "/api/v1/movie-title-match/batch/job-done/download"
    )


def test_download_expired_job_returns_410(_stub_batch_storage):
    import io as _io

    import openpyxl

    wb = openpyxl.Workbook()
    wb.active.append(["movie_title"])
    buf = _io.BytesIO()
    wb.save(buf)
    output_key = "batch-outputs/job-expired_output.xlsx"
    _stub_batch_storage[output_key] = buf.getvalue()

    with Session(_sqlite_engine) as session:
        job = MovieTitleBatchJob(
            id="job-expired",
            status="completed",
            total=1,
            processed=1,
            output_path=output_key,
            ttl=datetime.utcnow() - timedelta(hours=1),
        )
        session.add(job)
        session.commit()

    resp = client.get("/api/v1/movie-title-match/batch/job-expired/download")
    assert resp.status_code == 410


def test_download_missing_job_returns_404():
    resp = client.get("/api/v1/movie-title-match/batch/no-such-job/download")
    assert resp.status_code == 404


def test_download_completed_but_file_missing_returns_404():
    with Session(_sqlite_engine) as session:
        job = MovieTitleBatchJob(
            id="job-nofile",
            status="completed",
            total=1,
            processed=1,
            output_path="/tmp/movie_title_batch_outputs/does-not-exist.xlsx",
            ttl=datetime.utcnow() + timedelta(hours=1),
        )
        session.add(job)
        session.commit()

    resp = client.get("/api/v1/movie-title-match/batch/job-nofile/download")
    assert resp.status_code == 404


# ─────────────────────────────────────────────────────────────────────────────
# POST /batch — market="international"
# ─────────────────────────────────────────────────────────────────────────────

def test_domestic_upload_unaffected_when_market_omitted():
    """Regression guard: omitting market entirely must behave exactly like
    before market/country existed — MovieTitleBatchJob, not the intl table."""
    resp = client.post(
        "/api/v1/movie-title-match/batch",
        files={"file": ("titles.csv", _valid_csv(), "text/csv")},
    )
    assert resp.status_code == 200
    job_id = resp.json()["job_id"]
    assert _get_job(job_id) is not None
    assert _get_intl_job(job_id) is None


def test_international_upload_returns_200_and_creates_intl_job():
    resp = client.post(
        "/api/v1/movie-title-match/batch",
        files={"file": ("titles.csv", _valid_intl_csv(), "text/csv")},
        data={"market": "international"},
    )
    assert resp.status_code == 200
    job_id = resp.json()["job_id"]

    intl_job = _get_intl_job(job_id)
    assert intl_job is not None
    assert intl_job.total == 1
    # Must NOT also land in the domestic table
    assert _get_job(job_id) is None


def test_international_upload_missing_country_column_returns_400():
    bad_csv = _csv_bytes(
        "movie_title,show_date,ticketing_url",
        "Blue Beetle,2023-08-16,https://example.com/blue-beetle",
    )
    resp = client.post(
        "/api/v1/movie-title-match/batch",
        files={"file": ("titles.csv", bad_csv, "text/csv")},
        data={"market": "international"},
    )
    assert resp.status_code == 400
    assert "country" in resp.json()["detail"]


def test_international_upload_dispatches_intl_task_not_domestic(
    _stub_dispatch_intl_batch, _stub_dispatch_batch,
):
    resp = client.post(
        "/api/v1/movie-title-match/batch",
        files={"file": ("titles.csv", _valid_intl_csv(), "text/csv")},
        data={"market": "international"},
    )
    job_id = resp.json()["job_id"]
    assert _stub_dispatch_intl_batch == [job_id]
    assert _stub_dispatch_batch == []


# ─────────────────────────────────────────────────────────────────────────────
# GET /batch/intl/{job_id}, GET /batch/intl/{job_id}/download
# ─────────────────────────────────────────────────────────────────────────────

def test_intl_status_unknown_job_returns_404():
    resp = client.get("/api/v1/movie-title-match/batch/intl/does-not-exist")
    assert resp.status_code == 404


def test_intl_status_progress_computed_when_total_positive():
    with Session(_sqlite_engine) as session:
        job = MovieTitleIntlBatchJob(id="intl-job-progress", status="processing", total=4, processed=1)
        session.add(job)
        session.commit()

    resp = client.get("/api/v1/movie-title-match/batch/intl/intl-job-progress")
    assert resp.status_code == 200
    assert resp.json()["progress"] == 0.25


def test_intl_download_before_completion_returns_400():
    with Session(_sqlite_engine) as session:
        job = MovieTitleIntlBatchJob(id="intl-job-notdone", status="processing", total=1)
        session.add(job)
        session.commit()

    resp = client.get("/api/v1/movie-title-match/batch/intl/intl-job-notdone/download")
    assert resp.status_code == 400


def test_intl_download_after_completion_returns_200_xlsx_and_status_shows_intl_url(_stub_batch_storage):
    import io as _io

    import openpyxl

    wb = openpyxl.Workbook()
    wb.active.append(["movie_title", "country", "mapped_title"])
    buf = _io.BytesIO()
    wb.save(buf)
    output_key = "batch-outputs/intl-job-done_output.xlsx"
    _stub_batch_storage[output_key] = buf.getvalue()

    with Session(_sqlite_engine) as session:
        job = MovieTitleIntlBatchJob(
            id="intl-job-done",
            status="completed",
            total=1,
            processed=1,
            matched=1,
            output_path=output_key,
            ttl=datetime.utcnow() + timedelta(hours=1),
        )
        session.add(job)
        session.commit()

    resp = client.get("/api/v1/movie-title-match/batch/intl/intl-job-done/download")
    assert resp.status_code == 200
    assert resp.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    status_resp = client.get("/api/v1/movie-title-match/batch/intl/intl-job-done")
    assert status_resp.json()["output_url"] == (
        "/api/v1/movie-title-match/batch/intl/intl-job-done/download"
    )


def test_intl_download_missing_job_returns_404():
    resp = client.get("/api/v1/movie-title-match/batch/intl/no-such-job/download")
    assert resp.status_code == 404
