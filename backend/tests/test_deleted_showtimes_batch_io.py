"""Unit tests for app.deleted_showtimes.batch_io — pure helpers, no DB/network."""

import io

import openpyxl
import pytest

from app.deleted_showtimes.batch_io import (
    REQUIRED_COLUMNS,
    VERDICT_COL,
    build_output_xlsx,
    parse_upload,
    peek_headers,
    rows_to_showtime_rows,
)
from app.deleted_showtimes.core import FALSE_, TRUE_


def _csv_bytes(header_line: str, *data_lines: str) -> bytes:
    text = "\n".join([header_line, *data_lines]) + "\n"
    return text.encode("utf-8-sig")


def _xlsx_bytes(headers: list[str], rows: list[list]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class TestParseUpload:
    def test_parses_csv_with_required_columns(self):
        contents = _csv_bytes(
            "Theater Name,Title,Show date,Show time",
            "AMC Wayne 14,Dune,2026-08-06,7:00 PM",
        )
        headers, rows = parse_upload(contents, ".csv")
        assert headers == ["Theater Name", "Title", "Show date", "Show time"]
        assert len(rows) == 1
        assert rows[0]["Theater Name"] == "AMC Wayne 14"

    def test_missing_required_column_raises(self):
        contents = _csv_bytes("Theater Name,Title,Show date", "AMC Wayne 14,Dune,2026-08-06")
        with pytest.raises(ValueError, match="Show time"):
            parse_upload(contents, ".csv")

    def test_required_columns_matched_case_insensitively(self):
        contents = _csv_bytes(
            "theater name,title,show date,show time",
            "AMC Wayne 14,Dune,2026-08-06,7:00 PM",
        )
        headers, rows = parse_upload(contents, ".csv")
        assert len(rows) == 1

    def test_extra_columns_are_preserved(self):
        contents = _csv_bytes(
            "MM ID,Theater Name,Title,Show date,Show time,Total Seats",
            "123,AMC Wayne 14,Dune,2026-08-06,7:00 PM,200",
        )
        headers, rows = parse_upload(contents, ".csv")
        assert "MM ID" in headers
        assert rows[0]["Total Seats"] == "200"

    def test_xlsx_parses_with_required_columns(self):
        contents = _xlsx_bytes(
            ["Theater Name", "Title", "Show date", "Show time"],
            [["AMC Wayne 14", "Dune", "2026-08-06", "7:00 PM"]],
        )
        headers, rows = parse_upload(contents, ".xlsx")
        assert len(rows) == 1

    def test_unsupported_extension_raises(self):
        with pytest.raises(ValueError, match="Unsupported"):
            parse_upload(b"", ".txt")


class TestPeekHeaders:
    def test_peek_headers_csv(self):
        contents = _csv_bytes("Theater Name,Title,Show date,Show time")
        assert peek_headers(contents, ".csv") == ["Theater Name", "Title", "Show date", "Show time"]


class TestRowsToShowtimeRows:
    def test_builds_showtime_rows_with_parsed_time_and_date(self):
        headers = ["Theater Name", "Title", "Show date", "Show time"]
        rows = [{"Theater Name": "AMC Wayne 14", "Title": "Dune", "Show date": "2026-08-06", "Show time": "7:00 PM"}]

        srows = rows_to_showtime_rows(headers, rows)

        assert len(srows) == 1
        assert srows[0].theater == "AMC Wayne 14"
        assert srows[0].title == "Dune"
        assert srows[0].show_min == 19 * 60

    def test_missing_theater_or_unparseable_time_yields_none_fields(self):
        headers = ["Theater Name", "Title", "Show date", "Show time"]
        rows = [{"Theater Name": "", "Title": "Dune", "Show date": "2026-08-06", "Show time": "not a time"}]

        srows = rows_to_showtime_rows(headers, rows)

        assert srows[0].theater == ""
        assert srows[0].show_min is None


class TestBuildOutputXlsx:
    def test_appends_verdict_column_and_evidence_sheet(self):
        from app.deleted_showtimes.core import ShowtimeRow
        from datetime import date

        original_headers = ["Theater Name", "Title", "Show date", "Show time"]
        rows = [
            {"Theater Name": "AMC Wayne 14", "Title": "Dune", "Show date": "2026-08-06", "Show time": "7:00 PM"},
            {"Theater Name": "AMC Wayne 14", "Title": "Barbie", "Show date": "2026-08-06", "Show time": "8:00 PM"},
        ]
        showtime_rows = [
            ShowtimeRow(key=0, theater="AMC Wayne 14", title="Dune", show_date=date(2026, 8, 6),
                        show_time_raw="7:00 PM", show_min=1140, verdict=FALSE_, reason="EXACT_MATCH"),
            ShowtimeRow(key=1, theater="AMC Wayne 14", title="Barbie", show_date=date(2026, 8, 6),
                        show_time_raw="8:00 PM", show_min=1200, verdict=TRUE_, reason="NO_EXACT_MATCH"),
        ]

        xlsx_bytes = build_output_xlsx(original_headers, rows, showtime_rows)

        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
        assert "SERP_EVIDENCE" in wb.sheetnames

        ws = wb[wb.sheetnames[0]]
        rows_out = list(ws.iter_rows(values_only=True))
        assert rows_out[0] == ("Theater Name", "Title", "Show date", "Show time", VERDICT_COL)
        assert rows_out[1][-1] == FALSE_
        assert rows_out[2][-1] == TRUE_

        ev = wb["SERP_EVIDENCE"]
        ev_rows = list(ev.iter_rows(values_only=True))
        assert ev_rows[0][0] == "Row"
        assert ev_rows[1][5] == FALSE_  # DELETED_SHOWTIME column
        assert ev_rows[2][5] == TRUE_


def test_required_columns_constant():
    assert REQUIRED_COLUMNS == ("Theater Name", "Title", "Show date", "Show time")
