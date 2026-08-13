"""
Tests for app.workers.intl_batch_worker.run_intl_batch_job.

Calls run_intl_batch_job directly and synchronously (no thread) against a
real in-memory sqlite job row, exactly the way a worker function should be
tested. There is no S3 round-trip to cover here — the amenity workers are
purely local-filesystem (/tmp/...); S3 belongs to the deleted-showtimes
worker, not this one.
"""

from __future__ import annotations

import io
import json
import os

import openpyxl
import pytest
from sqlmodel import Session, SQLModel, create_engine, select
from sqlalchemy.pool import StaticPool

import app.workers.intl_batch_worker as worker_mod
from app.models import IntlDetectionJob
from app.intl_detection.engine import IntlMappingIndex, IntlScreenFormatEngine
from app.intl_detection.types import IntlApprovedMapping
from app.detection.normalizer import normalize_string, track_a_clean, track_b_clean, track_c_tokens


def _make_mapping(keyword: str, fmt: str, tier: int) -> IntlApprovedMapping:
    return IntlApprovedMapping(
        amenity_keyword=keyword,
        screen_format=fmt,
        priority_tier=tier,
        circuit_name=None,
        na_default=None,
        norm_exact=normalize_string(keyword).lower(),
        norm_track_a=track_a_clean(keyword),
        norm_track_b=track_b_clean(keyword),
        norm_track_c=track_c_tokens(keyword),
    )


@pytest.fixture
def engine():
    mappings = [
        _make_mapping("4DX", "4DX", 1),
        _make_mapping("ScreenX", "ScreenX", 3),
    ]
    return IntlScreenFormatEngine(IntlMappingIndex(mappings))


@pytest.fixture
def db_engine(monkeypatch):
    db = create_engine(
        "sqlite://",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    SQLModel.metadata.create_all(db)
    monkeypatch.setattr(worker_mod, "db_engine", db)
    yield db
    db.dispose()


@pytest.fixture
def job(db_engine):
    with Session(db_engine) as session:
        j = IntlDetectionJob(id="test-job-1", status="queued", total=0)
        session.add(j)
        session.commit()
    return "test-job-1"


def _get_job(db_engine, job_id) -> IntlDetectionJob:
    with Session(db_engine) as session:
        return session.get(IntlDetectionJob, job_id)


def _write_upload(tmp_path, headers, rows, ext=".xlsx"):
    if ext == ".xlsx":
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(headers)
        for row in rows:
            ws.append(row)
        path = tmp_path / f"upload{ext}"
        wb.save(path)
        return str(path)

    path = tmp_path / f"upload{ext}"
    lines = [",".join(headers)] + [",".join(str(c) for c in row) for row in rows]
    path.write_text("\n".join(lines) + "\n", encoding="utf-8-sig")
    return str(path)


def test_job_lifecycle_queued_to_completed(db_engine, job, engine, tmp_path):
    upload_path = _write_upload(tmp_path, ["amenities"], [["4DX"], ["ScreenX"], ["Comfy Recliners"]])

    worker_mod.run_intl_batch_job(job, upload_path, include_diagnostics=False, detection_engine=engine)

    j = _get_job(db_engine, job)
    assert j.status == "completed"
    assert j.processed == j.total == 3
    assert os.path.exists(j.output_path)


def test_output_workbook_has_appended_columns_in_order(db_engine, job, engine, tmp_path):
    upload_path = _write_upload(tmp_path, ["amenities", "circuit_name"], [["4DX", "AMC"]])

    worker_mod.run_intl_batch_job(job, upload_path, include_diagnostics=False, detection_engine=engine)

    j = _get_job(db_engine, job)
    wb = openpyxl.load_workbook(j.output_path)
    ws = wb.active
    header_row = [c.value for c in ws[1]]
    assert header_row == [
        "amenities",
        "circuit_name",
        "screen_format",
        "match_track",
        "confidence",
        "matched_keyword",
        "priority_tier",
        "match_source",
    ]


def test_csv_input_produces_xlsx_output(db_engine, job, engine, tmp_path):
    upload_path = _write_upload(tmp_path, ["amenities"], [["4DX"]], ext=".csv")

    worker_mod.run_intl_batch_job(job, upload_path, include_diagnostics=False, detection_engine=engine)

    j = _get_job(db_engine, job)
    assert j.output_path.endswith(".xlsx")
    # Confirm it really is a readable xlsx workbook, not a renamed csv.
    openpyxl.load_workbook(j.output_path)


def test_matched_and_no_match_counts_land_in_stats(db_engine, job, engine, tmp_path):
    upload_path = _write_upload(
        tmp_path, ["amenities"], [["4DX"], ["ScreenX"], ["Comfy Recliners"], ["Nonsense"]]
    )

    worker_mod.run_intl_batch_job(job, upload_path, include_diagnostics=False, detection_engine=engine)

    j = _get_job(db_engine, job)
    stats = json.loads(j.stats)
    assert stats["matched"] == 2
    assert stats["no_match"] == 2


def test_unmatched_row_gets_standard_and_no_match(db_engine, job, engine, tmp_path):
    upload_path = _write_upload(tmp_path, ["amenities"], [["Comfy Recliners"]])

    worker_mod.run_intl_batch_job(job, upload_path, include_diagnostics=False, detection_engine=engine)

    j = _get_job(db_engine, job)
    wb = openpyxl.load_workbook(j.output_path)
    ws = wb.active
    data_row = [c.value for c in ws[2]]
    header_row = [c.value for c in ws[1]]
    row_dict = dict(zip(header_row, data_row))
    assert row_dict["screen_format"] == "Standard"
    assert row_dict["match_source"] == "No Match"


def test_row_with_empty_amenity_cell_does_not_crash_the_job(db_engine, job, engine, tmp_path):
    upload_path = _write_upload(tmp_path, ["amenities"], [[""], ["4DX"]])

    worker_mod.run_intl_batch_job(job, upload_path, include_diagnostics=False, detection_engine=engine)

    j = _get_job(db_engine, job)
    assert j.status == "completed"
    assert j.processed == 2


def test_malformed_upload_marks_job_failed_not_hung(db_engine, job, engine, tmp_path):
    nonexistent_path = str(tmp_path / "does-not-exist.xlsx")

    # Should not raise — the worker's own try/except must catch this.
    worker_mod.run_intl_batch_job(job, nonexistent_path, include_diagnostics=False, detection_engine=engine)

    j = _get_job(db_engine, job)
    assert j.status == "failed"


def test_upload_file_is_deleted_after_completion(db_engine, job, engine, tmp_path):
    upload_path = _write_upload(tmp_path, ["amenities"], [["4DX"]])
    assert os.path.exists(upload_path)

    worker_mod.run_intl_batch_job(job, upload_path, include_diagnostics=False, detection_engine=engine)

    assert not os.path.exists(upload_path)


def test_ttl_is_set_on_completion(db_engine, job, engine, tmp_path):
    upload_path = _write_upload(tmp_path, ["amenities"], [["4DX"]])

    worker_mod.run_intl_batch_job(job, upload_path, include_diagnostics=False, detection_engine=engine)

    j = _get_job(db_engine, job)
    assert j.ttl is not None


def test_audit_mode_appends_match_status_column(db_engine, job, engine, tmp_path):
    upload_path = _write_upload(
        tmp_path,
        ["amenities", "screen_format"],
        [["4DX", "4DX"], ["4DX", "IMAX"]],
    )

    worker_mod.run_intl_batch_job(
        job, upload_path, include_diagnostics=False, detection_engine=engine, audit_mode=True
    )

    j = _get_job(db_engine, job)
    wb = openpyxl.load_workbook(j.output_path)
    ws = wb.active
    header_row = [c.value for c in ws[1]]
    assert header_row[-1] == "match_status"

    row1 = dict(zip(header_row, [c.value for c in ws[2]]))
    row2 = dict(zip(header_row, [c.value for c in ws[3]]))
    assert row1["match_status"] == "MATCH"
    assert row2["match_status"] == "MISMATCH"

    stats = json.loads(j.stats)
    assert stats["mismatch_count"] == 1
